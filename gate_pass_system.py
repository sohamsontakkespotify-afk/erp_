import sys, os, time, pickle
from datetime import datetime, date
import cv2
import numpy as np
import face_recognition
import face_recognition_models
from openpyxl import Workbook, load_workbook
import threading
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from PIL import Image, ImageTk
import queue

# Sound imports with fallback
try:
    import winsound  # For Windows
    SOUND_AVAILABLE = True
    SOUND_TYPE = "windows"
except ImportError:
    try:
        import pygame  # Cross-platform alternative
        pygame.mixer.init()
        SOUND_AVAILABLE = True
        SOUND_TYPE = "pygame"
    except ImportError:
        try:
            import playsound  # Another cross-platform option
            SOUND_AVAILABLE = True
            SOUND_TYPE = "playsound"
        except ImportError:
            SOUND_AVAILABLE = False
            SOUND_TYPE = None

EXCEL_FILE = "gatepass_logs.xlsx"
GATE_SHEET = "Gate_Entries"
GOING_OUT_SHEET = "Going_Out_Logs"

# ---------------------------
# Sound Functions
# ---------------------------
def play_sound(sound_type: str):
    """Play different sounds for entry/exit events"""
    if not SOUND_AVAILABLE:
        return
    
    def play_async():
        try:
            if SOUND_TYPE == "windows":
                if sound_type == "entry":
                    winsound.Beep(1000, 200)
                elif sound_type == "exit":
                    winsound.Beep(800, 150)
                    time.sleep(0.1)
                    winsound.Beep(800, 150)
                elif sound_type == "blocked":
                    for _ in range(3):
                        winsound.Beep(400, 100)
                        time.sleep(0.05)
                elif sound_type == "going_out":
                    winsound.Beep(600, 150)
                elif sound_type == "coming_back":
                    winsound.Beep(900, 200)
            
            elif SOUND_TYPE == "pygame":
                sample_rate = 22050
                if sound_type == "entry":
                    duration = 0.2
                    frequency = 1000
                    frames = int(duration * sample_rate)
                    arr = np.sin(2 * np.pi * frequency * np.linspace(0, duration, frames))
                    arr = (arr * 32767).astype(np.int16)
                    sound = pygame.sndarray.make_sound(np.array([arr, arr]).T)
                    sound.play()
                    
                elif sound_type == "exit":
                    duration = 0.15
                    frequency = 800
                    frames = int(duration * sample_rate)
                    arr = np.sin(2 * np.pi * frequency * np.linspace(0, duration, frames))
                    arr = (arr * 32767).astype(np.int16)
                    sound = pygame.sndarray.make_sound(np.array([arr, arr]).T)
                    sound.play()
                    time.sleep(0.1)
                    sound.play()
                
                elif sound_type == "blocked":
                    duration = 0.1
                    frequency = 400
                    frames = int(duration * sample_rate)
                    arr = np.sin(2 * np.pi * frequency * np.linspace(0, duration, frames))
                    arr = (arr * 32767).astype(np.int16)
                    sound = pygame.sndarray.make_sound(np.array([arr, arr]).T)
                    for _ in range(3):
                        sound.play()
                        time.sleep(0.05)
                        
                elif sound_type == "going_out":
                    duration = 0.15
                    frequency = 600
                    frames = int(duration * sample_rate)
                    arr = np.sin(2 * np.pi * frequency * np.linspace(0, duration, frames))
                    arr = (arr * 32767).astype(np.int16)
                    sound = pygame.sndarray.make_sound(np.array([arr, arr]).T)
                    sound.play()
                    
                elif sound_type == "coming_back":
                    duration = 0.2
                    frequency = 900
                    frames = int(duration * sample_rate)
                    arr = np.sin(2 * np.pi * frequency * np.linspace(0, duration, frames))
                    arr = (arr * 32767).astype(np.int16)
                    sound = pygame.sndarray.make_sound(np.array([arr, arr]).T)
                    sound.play()
                        
        except Exception as e:
            print(f"Sound error: {e}")
    
    threading.Thread(target=play_async, daemon=True).start()

# ---------------------------
# Excel Setup Functions
# ---------------------------
def setup_excel_file():
    """Initialize Excel file with both sheets"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        
        # Gate entries sheet (rename default sheet)
        ws_gate = wb.active
        ws_gate.title = GATE_SHEET
        ws_gate.append(["Name", "Phone_Number", "Entry_Time", "Exit_Time", "Status", "WhatsApp"])
        
        # Going out sheet
        ws_going_out = wb.create_sheet(GOING_OUT_SHEET)
        ws_going_out.append(["Name", "Phone_Number", "Going_Out_Time", "Coming_Back_Time", "Reason_Type", "Reason_Details", "Status", "Duration_Minutes"])
        
        wb.save(EXCEL_FILE)
    else:
        # Check if going out sheet exists, create if not
        wb = load_workbook(EXCEL_FILE)
        if GOING_OUT_SHEET not in wb.sheetnames:
            ws_going_out = wb.create_sheet(GOING_OUT_SHEET)
            ws_going_out.append(["Name", "Phone_Number", "Going_Out_Time", "Coming_Back_Time", "Reason_Type", "Reason_Details", "Status", "Duration_Minutes"])
            wb.save(EXCEL_FILE)
        wb.close()

# ---------------------------
# Excel Logging Functions
# ---------------------------
def get_user_status_and_history(name: str) -> tuple:
    """Check user status and return (status, has_exited_today, last_action_time)"""
    setup_excel_file()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb[GATE_SHEET] if GATE_SHEET in wb.sheetnames else wb.active
    
    today = date.today().strftime("%Y-%m-%d")
    
    today_entries = []
    last_entry = None
    last_action_time = None
    
    for row in ws.iter_rows(min_row=2):
        if row[0].value == name:
            last_entry = row
            
            if row[2].value:  # Entry_Time is now column 2 (index 2)
                try:
                    entry_date = datetime.strptime(str(row[2].value), "%Y-%m-%d %H:%M:%S").date()
                    if entry_date.strftime("%Y-%m-%d") == today:
                        today_entries.append(row)
                        # Get the most recent action time (entry or exit)
                        entry_time = datetime.strptime(str(row[2].value), "%Y-%m-%d %H:%M:%S")
                        if row[3].value:  # Exit_Time is now column 3 (index 3)
                            exit_time = datetime.strptime(str(row[3].value), "%Y-%m-%d %H:%M:%S")
                            last_action_time = max(entry_time, exit_time)
                        else:
                            last_action_time = entry_time
                except (ValueError, TypeError):
                    continue
    
    has_exited_today = False
    currently_inside = False
    
    for entry in today_entries:
        if entry[3].value is not None:  # Exit_Time is now column 3
            has_exited_today = True
        else:
            currently_inside = True
    
    wb.close()
    
    if currently_inside:
        return "INSIDE", has_exited_today, last_action_time
    elif has_exited_today:
        return "EXITED_TODAY", has_exited_today, last_action_time
    elif last_entry and last_entry[3].value is None:  # Exit_Time is now column 3
        return "INSIDE", has_exited_today, last_action_time
    else:
        return "OUTSIDE", has_exited_today, last_action_time

def get_going_out_status(name: str) -> tuple:
    """Check if user is currently out for work/personal reasons"""
    setup_excel_file()
    
    wb = load_workbook(EXCEL_FILE)
    ws_going_out = wb[GOING_OUT_SHEET]
    
    today = date.today().strftime("%Y-%m-%d")
    
    # Check for open going out entries today
    for row in ws_going_out.iter_rows(min_row=2):
        if row[0].value == name and row[2].value:  # Going_Out_Time is now column 2
            try:
                going_out_date = datetime.strptime(str(row[2].value), "%Y-%m-%d %H:%M:%S").date()
                if going_out_date.strftime("%Y-%m-%d") == today and not row[3].value:  # Coming_Back_Time is now column 3
                    # User is currently out
                    reason_type = row[4].value or "Unknown"  # Reason_Type is now column 4
                    going_out_time = datetime.strptime(str(row[2].value), "%Y-%m-%d %H:%M:%S")
                    wb.close()
                    return "OUT", reason_type, going_out_time
            except (ValueError, TypeError):
                continue
    
    wb.close()
    return "IN_OFFICE", None, None

def log_event_excel(name: str, override_cooling=False):
    """Log gate entry/exit events"""
    setup_excel_file()
    now = datetime.now()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb[GATE_SHEET] if GATE_SHEET in wb.sheetnames else wb.active
    
    # Remove old override column handling - no longer needed
    
    user_status, has_exited_today, last_action_time = get_user_status_and_history(name)
    
    # Check cooling period (2 minutes = 120 seconds) - can be overridden
    COOLING_PERIOD_SECONDS = 120
    
    if not override_cooling and last_action_time:
        time_since_last_action = (now - last_action_time).total_seconds()
        if time_since_last_action < COOLING_PERIOD_SECONDS:
            remaining_time = int(COOLING_PERIOD_SECONDS - time_since_last_action)
            minutes = remaining_time // 60
            seconds = remaining_time % 60
            status = f"COOLING_{minutes}m{seconds}s"
            play_sound("blocked")
            wb.close()
            return status
    
    phone = get_user_phone(name)
    today = now.strftime("%Y-%m-%d")
    
    # Find existing entry for today
    existing_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == name and row[2].value:
            # Handle both datetime objects and string formats
            entry_date = row[2].value
            if isinstance(entry_date, str):
                try:
                    entry_date = datetime.strptime(entry_date, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue
            
            if entry_date.strftime("%Y-%m-%d") == today:
                existing_row = row
                break
    
    if user_status == "OUTSIDE":
        # User is entering
        if existing_row and existing_row[3].value:
            # Already has exit time today, create new entry
            whatsapp_formula = '=HYPERLINK("https://web.whatsapp.com/send?phone=" & B' + str(ws.max_row + 1) + ' & "&text=" & ENCODEURL("Hi " & A' + str(ws.max_row + 1) + ' & ", your Entry time is " & TEXT(C' + str(ws.max_row + 1) + ',"yyyy-mm-dd hh:mm:ss") & IF(D' + str(ws.max_row + 1) + '<>"", " and Exit time is " & TEXT(D' + str(ws.max_row + 1) + ',"yyyy-mm-dd hh:mm:ss"), "") & " (Status: " & E' + str(ws.max_row + 1) + ')"), "Send WhatsApp")'
            ws.append([name, phone, now_str, None, "Inside", whatsapp_formula])
        elif existing_row:
            # Update existing row's entry time
            existing_row[2].value = now_str
            existing_row[4].value = "Inside"
            row_num = existing_row[0].row
            whatsapp_formula = f'=HYPERLINK("https://web.whatsapp.com/send?phone=" & B{row_num} & "&text=" & ENCODEURL("Hi " & A{row_num} & ", your Entry time is " & TEXT(C{row_num},"yyyy-mm-dd hh:mm:ss") & IF(D{row_num}<>"", " and Exit time is " & TEXT(D{row_num},"yyyy-mm-dd hh:mm:ss"), "") & " (Status: " & E{row_num} & ")"), "Send WhatsApp")'
            # Handle backward compatibility - add WhatsApp column if it doesn't exist
            if len(existing_row) > 5:
                existing_row[5].value = whatsapp_formula
            else:
                # Extend row to include WhatsApp column
                ws.cell(row=existing_row[0].row, column=6, value=whatsapp_formula)
        else:
            # Create new entry
            whatsapp_formula = '=HYPERLINK("https://web.whatsapp.com/send?phone=" & B' + str(ws.max_row + 1) + ' & "&text=" & ENCODEURL("Hi " & A' + str(ws.max_row + 1) + ' & ", your Entry time is " & TEXT(C' + str(ws.max_row + 1) + ',"yyyy-mm-dd hh:mm:ss") & IF(D' + str(ws.max_row + 1) + '<>"", " and Exit time is " & TEXT(D' + str(ws.max_row + 1) + ',"yyyy-mm-dd hh:mm:ss"), "") & " (Status: " & E' + str(ws.max_row + 1) + ')"), "Send WhatsApp")'
            ws.append([name, phone, now_str, None, "Inside", whatsapp_formula])
        status = "ENTRY_OVERRIDE" if override_cooling else "ENTRY"
        play_sound("entry")
    elif user_status == "EXITED_TODAY":
        status = "BLOCKED"
        play_sound("blocked")
        wb.close()
        return status
    else:
        # User is exiting - update existing entry with exit time
        if existing_row:
            existing_row[3].value = now_str
            existing_row[4].value = "Exited"
            row_num = existing_row[0].row
            whatsapp_formula = f'=HYPERLINK("https://web.whatsapp.com/send?phone=" & B{row_num} & "&text=" & ENCODEURL("Hi " & A{row_num} & ", your Entry time is " & TEXT(C{row_num},"yyyy-mm-dd hh:mm:ss") & IF(D{row_num}<>"", " and Exit time is " & TEXT(D{row_num},"yyyy-mm-dd hh:mm:ss"), "") & " (Status: " & E{row_num} & ")"), "Send WhatsApp")'
            # Handle backward compatibility - add WhatsApp column if it doesn't exist
            if len(existing_row) > 5:
                existing_row[5].value = whatsapp_formula
            else:
                # Extend row to include WhatsApp column
                ws.cell(row=existing_row[0].row, column=6, value=whatsapp_formula)
        else:
            # No entry found, create one with exit time only
            whatsapp_formula = '=HYPERLINK("https://web.whatsapp.com/send?phone=" & B' + str(ws.max_row + 1) + ' & "&text=" & ENCODEURL("Hi " & A' + str(ws.max_row + 1) + ' & ", your Exit time is " & TEXT(D' + str(ws.max_row + 1) + ',"yyyy-mm-dd hh:mm:ss") & " (Status: " & E' + str(ws.max_row + 1) + ')"), "Send WhatsApp")'
            ws.append([name, phone, None, now_str, "Exited", whatsapp_formula])
        status = "EXIT_OVERRIDE" if override_cooling else "EXIT"
        play_sound("exit")
    
    wb.save(EXCEL_FILE)
    wb.close()
    return status

def log_going_out_event(name: str, reason_type: str, reason_details: str = ""):
    """Log going out for work/personal reasons"""
    setup_excel_file()
    now = datetime.now()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")
    
    wb = load_workbook(EXCEL_FILE)
    ws_going_out = wb[GOING_OUT_SHEET]
    
    # Check current going out status
    going_out_status, current_reason, going_out_time = get_going_out_status(name)
    
    if going_out_status == "IN_OFFICE":
        # User is going out
        phone = get_user_phone(name)
        ws_going_out.append([name, phone, now_str, None, reason_type, reason_details, "Out", None])
        status = "GOING_OUT"
        play_sound("going_out")
    else:
        # User is coming back
        # Find the open going out entry
        found_open = None
        for row in reversed(list(ws_going_out.iter_rows(min_row=2))):
            if row[0].value == name and row[3].value is None:  # Coming_Back_Time is now column 3
                found_open = row
                break
        
        if found_open:
            found_open[3].value = now_str  # Coming back time (column 3)
            found_open[6].value = "Returned"  # Status (column 6)
            
            # Calculate duration in minutes
            if found_open[2].value:  # Going_Out_Time is column 2
                try:
                    going_out_dt = datetime.strptime(str(found_open[2].value), "%Y-%m-%d %H:%M:%S")
                    duration = (now - going_out_dt).total_seconds() / 60
                    found_open[7].value = round(duration, 1)  # Duration_Minutes is column 7
                except:
                    found_open[7].value = 0
            
            status = "COMING_BACK"
            play_sound("coming_back")
        else:
            # Shouldn't happen, but handle it
            phone = get_user_phone(name)
            ws_going_out.append([name, phone, None, now_str, reason_type, reason_details, "Returned", None])
            status = "COMING_BACK"
            play_sound("coming_back")
    
    wb.save(EXCEL_FILE)
    wb.close()
    return status

# ---------------------------
# Face Recognition Setup
# ---------------------------
def resource_path(relative_path: str) -> str:
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(relative_path)

def app_dir_path() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

MODELS_DIR = os.path.join("face_recognition_models", "models")
shape_predictor_68 = resource_path(os.path.join(MODELS_DIR, "shape_predictor_68_face_landmarks.dat"))
resnet_model = resource_path(os.path.join(MODELS_DIR, "dlib_face_recognition_resnet_model_v1.dat"))
face_recognition_models.pose_predictor_model_location = lambda: shape_predictor_68
face_recognition_models.face_recognition_model_location = lambda: resnet_model

PKL_PATH = os.path.join(app_dir_path(), "known_faces.pkl")
PHONE_DATA_PATH = os.path.join(app_dir_path(), "user_phones.pkl")
COOLDOWN_SECONDS = 3
MATCH_TOLERANCE = 0.5

def load_known_faces() -> dict:
    if not os.path.exists(PKL_PATH):
        return {}
    try:
        with open(PKL_PATH, "rb") as f:
            data = pickle.load(f)
        if isinstance(data, dict) and all(isinstance(v, np.ndarray) for v in data.values()):
            return data
        if isinstance(data, dict) and "encodings" in data and "names" in data:
            return {n: e for n, e in zip(data["names"], data["encodings"])}
        return {}
    except:
        return {}

def save_known_faces(known: dict) -> None:
    with open(PKL_PATH, "wb") as f:
        pickle.dump(known, f)

def load_user_phones() -> dict:
    """Load user phone numbers from file"""
    if not os.path.exists(PHONE_DATA_PATH):
        return {}
    try:
        with open(PHONE_DATA_PATH, "rb") as f:
            return pickle.load(f)
    except:
        return {}

def save_user_phones(phones: dict) -> None:
    """Save user phone numbers to file"""
    with open(PHONE_DATA_PATH, "wb") as f:
        pickle.dump(phones, f)

def get_user_phone(name: str) -> str:
    """Get phone number for a user"""
    phones = load_user_phones()
    return phones.get(name, "")

# ---------------------------
# GUI Application Class
# ---------------------------
class GatePassGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("ALANKAR - Gate Pass System")
        self.root.geometry("1100x750")
        self.root.configure(bg="#2c3e50")
        
        # Center the window
        self.root.eval('tk::PlaceWindow . center')
        
        # Variables
        self.is_scanning = False
        self.camera = None
        self.known_faces = {}
        self.last_scan = {}
        self.frame_queue = queue.Queue()
        
        # Colors
        self.colors = {
            'bg': '#2c3e50',
            'card': '#34495e',
            'primary': '#3498db',
            'success': '#27ae60',
            'warning': '#f39c12',
            'danger': '#e74c3c',
            'text': '#ecf0f1',
            'text_dark': '#2c3e50',
            'purple': '#8e44ad',
            'orange': '#e67e22'
        }
        
        self.setup_ui()
        self.load_faces()
        setup_excel_file()
        
    def setup_ui(self):
        # Title Bar
        title_frame = tk.Frame(self.root, bg=self.colors['primary'], height=80)
        title_frame.pack(fill='x')
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(
            title_frame,
            text="🚪 ALANKAR - Gate Pass System",
            font=("Arial", 24, "bold"),
            bg=self.colors['primary'],
            fg="white"
        )
        title_label.pack(expand=True)
        
        # Status Bar
        self.status_frame = tk.Frame(self.root, bg=self.colors['card'], height=60)
        self.status_frame.pack(fill='x', padx=10, pady=5)
        self.status_frame.pack_propagate(False)
        
        self.status_label = tk.Label(
            self.status_frame,
            text="🟢 System Ready",
            font=("Arial", 14),
            bg=self.colors['card'],
            fg=self.colors['text']
        )
        self.status_label.pack(side='left', padx=20, pady=15)
        
        self.datetime_label = tk.Label(
            self.status_frame,
            text="",
            font=("Arial", 12),
            bg=self.colors['card'],
            fg=self.colors['text']
        )
        self.datetime_label.pack(side='right', padx=20, pady=15)
        
        # Main Content
        main_frame = tk.Frame(self.root, bg=self.colors['bg'])
        main_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Left Panel - Camera/Recognition
        left_panel = tk.Frame(main_frame, bg=self.colors['card'], width=600)
        left_panel.pack(side='left', fill='both', expand=True, padx=(0, 5))
        left_panel.pack_propagate(False)
        
        # Camera Frame
        camera_label = tk.Label(
            left_panel,
            text="📹 Live Camera Feed",
            font=("Arial", 16, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text']
        )
        camera_label.pack(pady=10)
        
        self.video_frame = tk.Label(
            left_panel,
            text="Camera Off\n\nClick 'Start Recognition' to begin",
            font=("Arial", 14),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            width=580,
            height=380
        )
        self.video_frame.pack(padx=20, pady=(10, 5))
        
        # Message Display
        self.message_frame = tk.Frame(left_panel, bg=self.colors['card'])
        self.message_frame.pack(fill='x', padx=20, pady=10)
        
        self.message_label = tk.Label(
            self.message_frame,
            text="",
            font=("Arial", 14, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text'],
            height=2
        )
        self.message_label.pack()
        
        # Right Panel - Controls
        right_panel = tk.Frame(main_frame, bg=self.colors['card'], width=450)
        right_panel.pack(side='right', fill='y', padx=(5, 0))
        right_panel.pack_propagate(False)
        
        # Control Buttons
        controls_label = tk.Label(
            right_panel,
            text="⚙️ System Controls",
            font=("Arial", 16, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text']
        )
        controls_label.pack(pady=15)
        
        # Button style configuration
        btn_style = {
            'font': ("Arial", 11, "bold"),
            'width': 30,
            'height': 2,
            'fg': "white"
        }
        
        # Start Recognition Button
        self.start_btn = tk.Button(
            right_panel,
            text="🎥 Start Gate Recognition",
            bg=self.colors['success'],
            command=self.toggle_recognition,
            **btn_style
        )
        self.start_btn.pack(pady=5)
        
        # Separator for Going Out Section
        separator1 = tk.Frame(right_panel, height=2, bg=self.colors['text'])
        separator1.pack(fill='x', padx=20, pady=10)
        
        going_out_label = tk.Label(
            right_panel,
            text="🚶 Going Out Controls",
            font=("Arial", 14, "bold"),
            bg=self.colors['card'],
            fg=self.colors['orange']
        )
        going_out_label.pack(pady=(5, 10))
        
        # Going Out Button
        self.going_out_btn = tk.Button(
            right_panel,
            text="🚶‍♂️ Going Out (Work/Personal)",
            bg=self.colors['orange'],
            command=self.going_out_interface,
            font=("Arial", 11, "bold"),
            width=30,
            height=2,
            fg="white"
        )
        self.going_out_btn.pack(pady=5)
        
        # Separator for Emergency Section
        separator2 = tk.Frame(right_panel, height=2, bg=self.colors['text'])
        separator2.pack(fill='x', padx=20, pady=10)
        
        # Emergency Override Button
        self.override_btn = tk.Button(
            right_panel,
            text="🚨 Emergency Override",
            bg="#ff5722",  # Deep orange for emergency
            command=self.emergency_override,
            **btn_style
        )
        self.override_btn.pack(pady=5)
        
        # Separator for Management Section
        separator3 = tk.Frame(right_panel, height=2, bg=self.colors['text'])
        separator3.pack(fill='x', padx=20, pady=10)
        
        management_label = tk.Label(
            right_panel,
            text="👥 User Management",
            font=("Arial", 14, "bold"),
            bg=self.colors['card'],
            fg=self.colors['primary']
        )
        management_label.pack(pady=(5, 10))
        
        self.register_btn = tk.Button(
            right_panel,
            text="👤 Register New Person",
            bg=self.colors['primary'],
            command=self.register_person,
            **btn_style
        )
        self.register_btn.pack(pady=3)
        
        self.users_btn = tk.Button(
            right_panel,
            text="📋 View Registered Users",
            bg=self.colors['warning'],
            command=self.show_users,
            **btn_style
        )
        self.users_btn.pack(pady=3)
        
        self.delete_btn = tk.Button(
            right_panel,
            text="🗑️ Delete User",
            bg=self.colors['danger'],
            command=self.delete_user,
            **btn_style
        )
        self.delete_btn.pack(pady=3)
        
        # Separator for Logs Section
        separator4 = tk.Frame(right_panel, height=2, bg=self.colors['text'])
        separator4.pack(fill='x', padx=20, pady=10)
        
        logs_label = tk.Label(
            right_panel,
            text="📊 Reports & Logs",
            font=("Arial", 14, "bold"),
            bg=self.colors['card'],
            fg=self.colors['purple']
        )
        logs_label.pack(pady=(5, 10))
        
        self.logs_btn = tk.Button(
            right_panel,
            text="📊 View Gate Logs",
            bg=self.colors['purple'],
            command=self.view_logs,
            **btn_style
        )
        self.logs_btn.pack(pady=3)
        
        # New Going Out Logs Button
        self.going_out_logs_btn = tk.Button(
            right_panel,
            text="🚶 View Going Out Logs",
            bg="#d35400",  # Darker orange
            command=self.view_going_out_logs,
            **btn_style
        )
        self.going_out_logs_btn.pack(pady=3)
        
        # Exit Button
        exit_btn = tk.Button(
            right_panel,
            text="🚪 Exit System",
            font=("Arial", 12, "bold"),
            bg=self.colors['danger'],
            fg="white",
            width=30,
            height=2,
            command=self.close_application
        )
        exit_btn.pack(side='bottom', pady=20)
        
        # Start update loops
        self.update_datetime()
        
    def update_datetime(self):
        """Update datetime display"""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.datetime_label.config(text=f"🕐 {now}")
        self.root.after(1000, self.update_datetime)
    
    def load_faces(self):
        """Load known faces from file"""
        self.known_faces = load_known_faces()
        
    def toggle_recognition(self):
        """Start/stop face recognition"""
        if not self.is_scanning:
            self.start_recognition()
        else:
            self.stop_recognition()
            
    def start_recognition(self):
        """Start face recognition process"""
        if len(self.known_faces) == 0:
            messagebox.showwarning("No Users", "Please register at least one person before starting recognition.")
            return

        try:
            # Use DirectShow backend for faster startup (Windows only)
            self.camera = cv2.VideoCapture(0, cv2.CAP_DSHOW)

            if not self.camera.isOpened():
                messagebox.showerror("Camera Error", "Could not access camera. Please check if camera is connected.")
                return

            # Set resolution for faster performance
            self.camera.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            self.camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

            # Warm-up frames to reduce initial lag
            for _ in range(5):
                self.camera.read()

            self.is_scanning = True
            self.start_btn.config(
                text="🛑 Stop Gate Recognition",
                bg=self.colors['danger']
            )
            self.status_label.config(
                text="🔍 Scanning for faces...",
                fg=self.colors['success']
            )

            # Start camera thread
            self.camera_thread = threading.Thread(target=self.camera_loop, daemon=True)
            self.camera_thread.start()

            # Start frame processing
            self.process_frames()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to start recognition: {str(e)}")

            
    def stop_recognition(self):
        """Stop face recognition process"""
        self.is_scanning = False
        
        if self.camera:
            self.camera.release()
            self.camera = None
            
        self.start_btn.config(
            text="🎥 Start Gate Recognition",
            bg=self.colors['success']
        )
        self.status_label.config(
            text="🟢 System Ready",
            fg=self.colors['text']
        )
        
        # Clear video frame
        self.video_frame.config(
            image="",
            text="Camera Off\n\nClick 'Start Recognition' to begin"
        )
        
    def camera_loop(self):
        """Camera capture loop running in separate thread"""
        known_names = list(self.known_faces.keys())
        known_encodings = np.stack(list(self.known_faces.values())) if known_names else np.empty((0, 128))
        
        while self.is_scanning and self.camera and self.camera.isOpened():
            ret, frame = self.camera.read()
            if not ret:
                break
                
            # Resize for faster processing
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = small_frame[:, :, ::-1]
            
            # Find faces
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
            
            for face_encoding in face_encodings:
                if known_encodings.size > 0:
                    matches = face_recognition.compare_faces(known_encodings, face_encoding, tolerance=MATCH_TOLERANCE)
                    face_distances = face_recognition.face_distance(known_encodings, face_encoding)
                    
                    if len(face_distances) > 0:
                        best_match_index = np.argmin(face_distances)
                        
                        if matches[best_match_index]:
                            name = known_names[best_match_index]
                            current_time = time.time()
                            
                            # Check cooldown
                            if name not in self.last_scan or (current_time - self.last_scan[name]) > COOLDOWN_SECONDS:
                                status = log_event_excel(name)
                                self.last_scan[name] = current_time
                                
                                # Queue the result for UI update
                                self.frame_queue.put(('recognition', name, status))
            
            # Queue frame for display
            # Convert BGR to RGB for tkinter
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            self.frame_queue.put(('frame', rgb_frame, None))
            
            time.sleep(0.03)  # ~30 FPS
            
    def process_frames(self):
        """Process frames from camera thread and update UI"""
        try:
            while not self.frame_queue.empty():
                item = self.frame_queue.get_nowait()
                
                if item[0] == 'frame':
                    # Update video display
                    rgb_frame = item[1]
                    # Resize frame for display - match registration window size
                    display_frame = cv2.resize(rgb_frame, (640, 480))
                    
                    # Convert to PIL Image and then to PhotoImage
                    pil_image = Image.fromarray(display_frame)
                    photo = ImageTk.PhotoImage(image=pil_image)
                    
                    self.video_frame.config(image=photo, text="")
                    self.video_frame.image = photo  # Keep a reference
                    
                elif item[0] == 'recognition':
                    # Update recognition result
                    name, status = item[1], item[2]
                    self.show_recognition_result(name, status)
                    
        except queue.Empty:
            pass
        
        # Schedule next frame processing
        if self.is_scanning:
            self.root.after(30, self.process_frames)
            
    def show_recognition_result(self, name, status):
        """Show recognition result in UI"""
        current_time = datetime.now().strftime("%H:%M:%S")
        
        if status == "ENTRY":
            message = f"✅ GATE ENTRY: {name} ({current_time})"
            bg_color = self.colors['success']
        elif status == "ENTRY_OVERRIDE":
            message = f"🚨 OVERRIDE ENTRY: {name} ({current_time})"
            bg_color = "#ff5722"  # Emergency orange
        elif status == "EXIT":
            message = f"🚪 GATE EXIT: {name} ({current_time})"
            bg_color = self.colors['warning']
        elif status == "EXIT_OVERRIDE":
            message = f"🚨 OVERRIDE EXIT: {name} ({current_time})"
            bg_color = "#ff5722"  # Emergency orange
        elif status == "BLOCKED":
            message = f"🚫 BLOCKED: {name} - Already exited today"
            bg_color = self.colors['danger']
        elif status.startswith("COOLING_"):
            # Extract remaining time from status
            time_part = status.replace("COOLING_", "")
            message = f"⏳ COOLING: {name} - Wait {time_part} before next action"
            bg_color = "#ff9800"  # Orange color for cooling
        else:
            message = f"❓ Unknown status for {name}"
            bg_color = self.colors['card']
            
        self.message_label.config(
            text=message,
            bg=bg_color,
            fg="white" if status in ["ENTRY", "BLOCKED", "ENTRY_OVERRIDE", "EXIT_OVERRIDE"] or status.startswith("COOLING_") else self.colors['text_dark']
        )
        
        # Clear message after 5 seconds (8 seconds for overrides)
        clear_time = 8000 if "OVERRIDE" in status else 5000
        self.root.after(clear_time, lambda: self.message_label.config(text="", bg=self.colors['card']))
        
    def emergency_override(self):
        """Allow manual override of cooling period for urgent situations"""
        if not self.known_faces:
            messagebox.showinfo("No Users", "No users registered.")
            return
        
        # Create override window
        override_window = tk.Toplevel(self.root)
        override_window.title("Emergency Override")
        override_window.geometry("500x400")
        override_window.configure(bg=self.colors['bg'])
        override_window.grab_set()
        
        # Title
        title_label = tk.Label(
            override_window,
            text="🚨 Emergency Override",
            font=("Arial", 18, "bold"),
            bg=self.colors['bg'],
            fg="#ff5722"
        )
        title_label.pack(pady=20)
        
        # Warning message
        warning_label = tk.Label(
            override_window,
            text="⚠️ This will bypass the cooling period for legitimate work needs\n(Personal emergencies, official business, urgent deliveries, etc.)",
            font=("Arial", 12),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            justify='center'
        )
        warning_label.pack(pady=10)
        
        # User selection
        selection_frame = tk.Frame(override_window, bg=self.colors['card'])
        selection_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(
            selection_frame,
            text="Select User:",
            font=("Arial", 14, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text']
        ).pack(pady=10)
        
        # Listbox for users in cooling period
        listbox_frame = tk.Frame(selection_frame, bg=self.colors['card'])
        listbox_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(listbox_frame)
        scrollbar.pack(side='right', fill='y')
        
        user_listbox = tk.Listbox(
            listbox_frame,
            font=("Arial", 12),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            selectbackground=self.colors['primary'],
            yscrollcommand=scrollbar.set
        )
        user_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=user_listbox.yview)
        
        # Populate with users and their cooling status
        cooling_users = []
        now = datetime.now()
        
        for name in self.known_faces.keys():
            status, _, last_action_time = get_user_status_and_history(name)
            if last_action_time:
                time_since_last = (now - last_action_time).total_seconds()
                if time_since_last < 120:  # Still in cooling period
                    remaining = int(120 - time_since_last)
                    minutes = remaining // 60
                    seconds = remaining % 60
                    cooling_users.append(name)
                    user_listbox.insert(tk.END, f"⏳ {name} - {minutes}m {seconds}s remaining")
                else:
                    user_listbox.insert(tk.END, f"✅ {name} - No cooling period")
            else:
                user_listbox.insert(tk.END, f"✅ {name} - No recent activity")
        
        # Buttons
        button_frame = tk.Frame(override_window, bg=self.colors['bg'])
        button_frame.pack(pady=20)
        
        def confirm_override():
            selection = user_listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a user.")
                return
            
            selected_text = user_listbox.get(selection[0])
            # Extract name from the display text
            if " - " in selected_text:
                name = selected_text.split(" ")[1]  # Get name after emoji
            else:
                messagebox.showerror("Error", "Could not determine user name.")
                return
            
            # Confirm override
            reason = simpledialog.askstring(
                "Override Reason", 
                f"Override cooling period for {name}?\n\nPlease enter reason (optional):",
                initialvalue="Work-related emergency",
                parent=self.root
            )
            
            if reason is not None:  # User didn't cancel
                try:
                    status = log_event_excel(name, override_cooling=True)
                    
                    if status.startswith("ENTRY"):
                        msg = f"✅ OVERRIDE ENTRY: {name}"
                        bg_color = self.colors['success']
                    elif status.startswith("EXIT"):
                        msg = f"🚪 OVERRIDE EXIT: {name}"
                        bg_color = self.colors['warning']
                    else:
                        msg = f"❓ Override result: {status}"
                        bg_color = self.colors['card']
                    
                    # Update main window message
                    self.message_label.config(
                        text=f"{msg}\nReason: {reason}",
                        bg=bg_color,
                        fg="white" if status.startswith(("ENTRY", "EXIT")) else self.colors['text']
                    )
                    
                    messagebox.showinfo("Override Complete", f"Cooling period bypassed for {name}\nReason: {reason}")
                    override_window.destroy()
                    
                    # Clear message after 8 seconds (longer for override messages)
                    self.root.after(8000, lambda: self.message_label.config(text="", bg=self.colors['card']))
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Override failed: {str(e)}")
        
        override_btn = tk.Button(
            button_frame,
            text="🚨 Override Cooling Period",
            font=("Arial", 12, "bold"),
            bg="#ff5722",
            fg="white",
            width=25,
            command=confirm_override
        )
        override_btn.pack(side='left', padx=10)
        
        cancel_btn = tk.Button(
            button_frame,
            text="❌ Cancel",
            font=("Arial", 12, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text'],
            width=15,
            command=override_window.destroy
        )
        cancel_btn.pack(side='left', padx=10)
        
    def register_person(self):
        """Register a new person"""
        name = simpledialog.askstring("Register Person", "Enter person's name/ID:", parent=self.root)
        if not name or not name.strip():
            return
            
        name = name.strip()
        
        # Ask for phone number
        phone = simpledialog.askstring("Phone Number", f"Enter phone number for {name}:", parent=self.root)
        if not phone or not phone.strip():
            if not messagebox.askyesno("No Phone Number", "No phone number entered. Continue registration without phone number?"):
                return
            phone = ""
        else:
            phone = phone.strip()
        
        if name in self.known_faces:
            if not messagebox.askyesno("User Exists", f"'{name}' already exists. Do you want to update their face data and phone number?"):
                return
        
        # Stop recognition temporarily
        was_scanning = self.is_scanning
        if was_scanning:
            self.stop_recognition()
            
        self.register_face_capture(name, phone, was_scanning)
        
    def register_face_capture(self, name, phone, restart_scanning):
        """Capture face for registration"""
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Camera Error", "Could not access camera for registration.")
            return
            
        # Create registration window
        reg_window = tk.Toplevel(self.root)
        reg_window.title("Face Registration")
        reg_window.geometry("800x600")
        reg_window.configure(bg=self.colors['bg'])
        reg_window.grab_set()  # Make modal
        
        # Instructions
        instruction_label = tk.Label(
            reg_window,
            text=f"Registering: {name}\n\nLook directly at the camera and click 'Capture' when ready",
            font=("Arial", 14),
            bg=self.colors['bg'],
            fg=self.colors['text']
        )
        instruction_label.pack(pady=20)
        
        # Video display
        video_label = tk.Label(reg_window, bg=self.colors['bg'])
        video_label.pack(pady=10)
        
        # Buttons
        button_frame = tk.Frame(reg_window, bg=self.colors['bg'])
        button_frame.pack(pady=20)
        
        capture_btn = tk.Button(
            button_frame,
            text="📸 Capture",
            font=("Arial", 12, "bold"),
            bg=self.colors['success'],
            fg="white",
            width=15,
            command=lambda: self.capture_face(cap, name, phone, reg_window, restart_scanning)
        )
        capture_btn.pack(side='left', padx=10)
        
        cancel_btn = tk.Button(
            button_frame,
            text="❌ Cancel",
            font=("Arial", 12, "bold"),
            bg=self.colors['danger'],
            fg="white",
            width=15,
            command=lambda: self.cancel_registration(cap, reg_window, restart_scanning)
        )
        cancel_btn.pack(side='left', padx=10)
        
        # Start video feed
        self.update_registration_video(cap, video_label, reg_window)
        
    def update_registration_video(self, cap, video_label, window):
        """Update video feed in registration window"""
        if cap.isOpened() and window.winfo_exists():
            ret, frame = cap.read()
            if ret:
                # Convert and resize
                rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                display_frame = cv2.resize(rgb_frame, (640, 480))
                
                # Convert to PhotoImage
                pil_image = Image.fromarray(display_frame)
                photo = ImageTk.PhotoImage(image=pil_image)
                
                video_label.config(image=photo)
                video_label.image = photo
                
                # Schedule next update
                window.after(30, lambda: self.update_registration_video(cap, video_label, window))
                
    def capture_face(self, cap, name, phone, window, restart_scanning):
        """Capture and process face for registration"""
        ret, frame = cap.read()
        if not ret:
            messagebox.showerror("Error", "Failed to capture image from camera.")
            return
            
        # Process face
        rgb_frame = frame[:, :, ::-1]
        face_locations = face_recognition.face_locations(rgb_frame)
        
        if len(face_locations) != 1:
            messagebox.showerror(
                "Face Detection Error", 
                f"Need exactly one face (found {len(face_locations)}). Please try again."
            )
            return
            
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        if not face_encodings:
            messagebox.showerror("Error", "Could not extract face features. Please try again.")
            return
            
        # Save the face
        self.known_faces[name] = face_encodings[0]
        save_known_faces(self.known_faces)
        
        # Save the phone number with +91 prefix if not empty
        phones = load_user_phones()
        if phone and not phone.startswith('+91'):
            phone = '91' + phone
        phones[name] = phone
        save_user_phones(phones)
        
        # Cleanup
        cap.release()
        window.destroy()
        
        phone_msg = f" (Phone: {phone})" if phone else " (No phone number)"
        messagebox.showinfo("Success", f"Successfully registered '{name}'!{phone_msg}")
        
        # Restart scanning if it was running
        if restart_scanning:
            self.start_recognition()
            
    def cancel_registration(self, cap, window, restart_scanning):
        """Cancel registration process"""
        cap.release()
        window.destroy()
        
        # Restart scanning if it was running
        if restart_scanning:
            self.start_recognition()
            
    def show_users(self):
        """Show registered users window"""
        users_window = tk.Toplevel(self.root)
        users_window.title("Registered Users")
        users_window.geometry("700x600")
        users_window.configure(bg=self.colors['bg'])
        
        # Title
        title_label = tk.Label(
            users_window,
            text="👥 Registered Users",
            font=("Arial", 18, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['text']
        )
        title_label.pack(pady=20)
        
        # Users list frame
        list_frame = tk.Frame(users_window, bg=self.colors['card'])
        list_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Scrollable text widget
        text_frame = tk.Frame(list_frame)
        text_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side='right', fill='y')
        
        users_text = tk.Text(
            text_frame,
            font=("Arial", 12),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            yscrollcommand=scrollbar.set,
            state='disabled'
        )
        users_text.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=users_text.yview)
        
        # Populate users list
        users_text.config(state='normal')
        if not self.known_faces:
            users_text.insert(tk.END, "No users registered yet.\n")
        else:
            users_text.insert(tk.END, f"Total Users: {len(self.known_faces)}\n\n")
            users_text.insert(tk.END, "=" * 60 + "\n")
            
            for i, name in enumerate(self.known_faces.keys(), 1):
                gate_status, has_exited_today, last_action_time = get_user_status_and_history(name)
                going_out_status, reason_type, going_out_time = get_going_out_status(name)
                
                # Gate Status
                if gate_status == "INSIDE":
                    gate_emoji = "🟢"
                    gate_text = "In Office"
                elif gate_status == "EXITED_TODAY":
                    gate_emoji = "🔴"
                    gate_text = "Exited Today (Blocked)"
                else:
                    gate_emoji = "⚪"
                    gate_text = "Outside"
                
                # Going Out Status
                if going_out_status == "OUT":
                    going_out_emoji = "🚶"
                    duration = ""
                    if going_out_time:
                        minutes = int((datetime.now() - going_out_time).total_seconds() / 60)
                        duration = f" ({minutes}m)"
                    going_out_text = f"Out for {reason_type}{duration}"
                else:
                    going_out_emoji = "🏢"
                    going_out_text = "In Office"
                
                phone = get_user_phone(name)
                phone_display = f" (📞 {phone})" if phone else " (📞 No phone)"
                
                users_text.insert(tk.END, f"{i}. {name}{phone_display}\n")
                users_text.insert(tk.END, f"   Gate Status: {gate_emoji} {gate_text}\n")
                users_text.insert(tk.END, f"   Going Out: {going_out_emoji} {going_out_text}\n")
                
                # Show cooling period info if applicable
                if last_action_time:
                    now = datetime.now()
                    time_since_last = (now - last_action_time).total_seconds()
                    if time_since_last < 120:  # 2 minutes cooling period
                        remaining = int(120 - time_since_last)
                        minutes = remaining // 60
                        seconds = remaining % 60
                        users_text.insert(tk.END, f"   Cooling: ⏳ {minutes}m {seconds}s remaining\n")
                    else:
                        users_text.insert(tk.END, f"   Last Gate Action: {last_action_time.strftime('%H:%M:%S')}\n")
                else:
                    users_text.insert(tk.END, f"   {'Has exited today' if has_exited_today else 'No gate activity today'}\n")
                
                users_text.insert(tk.END, "-" * 50 + "\n")
                
        users_text.config(state='disabled')
        
        # Close button
        close_btn = tk.Button(
            users_window,
            text="Close",
            font=("Arial", 12, "bold"),
            bg=self.colors['primary'],
            fg="white",
            width=20,
            command=users_window.destroy
        )
        close_btn.pack(pady=20)
        
    def delete_user(self):
        """Delete a user"""
        if not self.known_faces:
            messagebox.showinfo("No Users", "No users registered to delete.")
            return
            
        # Create selection window
        delete_window = tk.Toplevel(self.root)
        delete_window.title("Delete User")
        delete_window.geometry("400x500")
        delete_window.configure(bg=self.colors['bg'])
        delete_window.grab_set()
        
        # Title
        title_label = tk.Label(
            delete_window,
            text="🗑️ Delete User",
            font=("Arial", 18, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['text']
        )
        title_label.pack(pady=20)
        
        # Instructions
        instruction_label = tk.Label(
            delete_window,
            text="Select a user to delete:",
            font=("Arial", 12),
            bg=self.colors['bg'],
            fg=self.colors['text']
        )
        instruction_label.pack(pady=10)
        
        # Listbox for users
        listbox_frame = tk.Frame(delete_window, bg=self.colors['card'])
        listbox_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        scrollbar = ttk.Scrollbar(listbox_frame)
        scrollbar.pack(side='right', fill='y')
        
        user_listbox = tk.Listbox(
            listbox_frame,
            font=("Arial", 12),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            selectbackground=self.colors['primary'],
            yscrollcommand=scrollbar.set
        )
        user_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=user_listbox.yview)
        
        # Populate listbox
        for name in self.known_faces.keys():
            gate_status, _, _ = get_user_status_and_history(name)
            status_emoji = "🟢" if gate_status == "INSIDE" else ("🔴" if gate_status == "EXITED_TODAY" else "⚪")
            user_listbox.insert(tk.END, f"{status_emoji} {name}")
        
        # Buttons
        button_frame = tk.Frame(delete_window, bg=self.colors['bg'])
        button_frame.pack(pady=20)
        
        def confirm_delete():
            selection = user_listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a user to delete.")
                return
                
            selected_text = user_listbox.get(selection[0])
            name = selected_text.split(" ", 1)[1]  # Remove emoji
            
            if messagebox.askyesno(
                "Confirm Delete", 
                f"Are you sure you want to delete '{name}'?\n\nThis action cannot be undone."
            ):
                del self.known_faces[name]
                save_known_faces(self.known_faces)
                messagebox.showinfo("Success", f"User '{name}' has been deleted.")
                delete_window.destroy()
        
        delete_btn = tk.Button(
            button_frame,
            text="🗑️ Delete Selected",
            font=("Arial", 12, "bold"),
            bg=self.colors['danger'],
            fg="white",
            width=20,
            command=confirm_delete
        )
        delete_btn.pack(side='left', padx=10)
        
        cancel_btn = tk.Button(
            button_frame,
            text="❌ Cancel",
            font=("Arial", 12, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text'],
            width=15,
            command=delete_window.destroy
        )
        cancel_btn.pack(side='left', padx=10)
        
    def view_logs(self):
        """View entry/exit logs"""
        setup_excel_file()
        
        logs_window = tk.Toplevel(self.root)
        logs_window.title("Gate Entry/Exit Logs")
        logs_window.geometry("900x600")
        logs_window.configure(bg=self.colors['bg'])
        
        # Title
        title_label = tk.Label(
            logs_window,
            text="📊 Gate Entry/Exit Logs",
            font=("Arial", 18, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['text']
        )
        title_label.pack(pady=20)
        
        # Filter frame
        filter_frame = tk.Frame(logs_window, bg=self.colors['card'])
        filter_frame.pack(fill='x', padx=20, pady=10)
        
        # Today's logs button
        today_btn = tk.Button(
            filter_frame,
            text="📅 Today's Logs",
            font=("Arial", 10, "bold"),
            bg=self.colors['primary'],
            fg="white",
            command=lambda: self.load_logs(logs_text, today_only=True)
        )
        today_btn.pack(side='left', padx=10, pady=10)
        
        # All logs button
        all_btn = tk.Button(
            filter_frame,
            text="📋 All Logs",
            font=("Arial", 10, "bold"),
            bg=self.colors['warning'],
            fg="white",
            command=lambda: self.load_logs(logs_text, today_only=False)
        )
        all_btn.pack(side='left', padx=10, pady=10)
        
        # Logs display frame
        logs_frame = tk.Frame(logs_window, bg=self.colors['card'])
        logs_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Scrollable text widget
        text_frame = tk.Frame(logs_frame)
        text_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        scrollbar_v = ttk.Scrollbar(text_frame)
        scrollbar_v.pack(side='right', fill='y')
        
        scrollbar_h = ttk.Scrollbar(text_frame, orient='horizontal')
        scrollbar_h.pack(side='bottom', fill='x')
        
        logs_text = tk.Text(
            text_frame,
            font=("Courier", 10),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            yscrollcommand=scrollbar_v.set,
            xscrollcommand=scrollbar_h.set,
            wrap='none'
        )
        logs_text.pack(side='left', fill='both', expand=True)
        
        scrollbar_v.config(command=logs_text.yview)
        scrollbar_h.config(command=logs_text.xview)
        
        # Load today's logs by default
        self.load_logs(logs_text, today_only=True)
        
        # Close button
        close_btn = tk.Button(
            logs_window,
            text="Close",
            font=("Arial", 12, "bold"),
            bg=self.colors['primary'],
            fg="white",
            width=20,
            command=logs_window.destroy
        )
        close_btn.pack(pady=20)
        
    def load_logs(self, text_widget, today_only=True):
        """Load and display gate logs"""
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb[GATE_SHEET] if GATE_SHEET in wb.sheetnames else wb.active
            
            text_widget.config(state='normal')
            text_widget.delete(1.0, tk.END)
            
            today = date.today().strftime("%Y-%m-%d")
            
            # Header
            if today_only:
                text_widget.insert(tk.END, f"📅 Today's Gate Logs ({today})\n\n")
            else:
                text_widget.insert(tk.END, "📋 All Gate Logs\n\n")
                
            text_widget.insert(tk.END, f"{'Name':<15} {'Phone':<12} {'Entry Time':<15} {'Exit Time':<15} {'Status':<8} {'Override':<8}\n")
            text_widget.insert(tk.END, "=" * 90 + "\n")
            
            row_count = 0
            for row in ws.iter_rows(min_row=2):
                name = row[0].value or ""
                phone = row[1].value or ""  # Phone_Number is now column 1
                entry_time = row[2].value or ""  # Entry_Time is now column 2
                exit_time = row[3].value or ""  # Exit_Time is now column 3
                status = row[4].value or ""  # Status is now column 4
                override_used = row[5].value if len(row) > 5 else ""  # Override_Used is now column 5
                
                # Filter for today if requested
                if today_only and entry_time:
                    try:
                        entry_date = datetime.strptime(str(entry_time), "%Y-%m-%d %H:%M:%S").date()
                        if entry_date.strftime("%Y-%m-%d") != today:
                            continue
                    except:
                        continue
                
                # Format times
                entry_display = str(entry_time)[11:19] if entry_time and len(str(entry_time)) > 19 else str(entry_time)
                exit_display = str(exit_time)[11:19] if exit_time and len(str(exit_time)) > 19 else str(exit_time)
                override_display = "YES" if override_used == "YES" else ""
                phone_display = phone[:12] if phone else "N/A"  # Truncate phone for display
                
                text_widget.insert(tk.END, f"{name:<15} {phone_display:<12} {entry_display:<15} {exit_display:<15} {status:<8} {override_display:<8}\n")
                row_count += 1
                
            if row_count == 0:
                text_widget.insert(tk.END, "\nNo records found.\n")
            else:
                text_widget.insert(tk.END, f"\nTotal Records: {row_count}\n")
                
            text_widget.config(state='disabled')
            wb.close()
            
        except Exception as e:
            text_widget.config(state='normal')
            text_widget.delete(1.0, tk.END)
            text_widget.insert(tk.END, f"Error loading logs: {str(e)}")
            text_widget.config(state='disabled')
    
    def close_application(self):
        """Close the application"""
        if messagebox.askyesno("Exit", "Are you sure you want to exit the Gate Pass System?"):
            # Stop recognition if running
            if self.is_scanning:
                self.stop_recognition()
            
            # Clean up
            cv2.destroyAllWindows()
            self.root.quit()
            self.root.destroy()
    
    def run(self):
        """Start the GUI application"""
        # Handle window close event
        self.root.protocol("WM_DELETE_WINDOW", self.close_application)
        
        # Start the main loop
        try:
            self.root.mainloop()
        except Exception as e:
            print(f"Error in main loop: {e}")
        finally:
            # Cleanup
            if self.is_scanning:
                self.stop_recognition()
            cv2.destroyAllWindows()

    def view_logs(self):
        """View entry/exit logs"""
        setup_excel_file()
        
        logs_window = tk.Toplevel(self.root)
        logs_window.title("Gate Entry/Exit Logs")
        logs_window.geometry("900x600")
        logs_window.configure(bg=self.colors['bg'])
        
        # Title
        title_label = tk.Label(
            logs_window,
            text="📊 Gate Entry/Exit Logs",
            font=("Arial", 18, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['text']
        )
        title_label.pack(pady=20)
        
        # Filter frame
        filter_frame = tk.Frame(logs_window, bg=self.colors['card'])
        filter_frame.pack(fill='x', padx=20, pady=10)
        
        # Today's logs button
        today_btn = tk.Button(
            filter_frame,
            text="📅 Today's Logs",
            font=("Arial", 10, "bold"),
            bg=self.colors['primary'],
            fg="white",
            command=lambda: self.load_logs(logs_text, today_only=True)
        )
        today_btn.pack(side='left', padx=10, pady=10)
        
        # All logs button
        all_btn = tk.Button(
            filter_frame,
            text="📋 All Logs",
            font=("Arial", 10, "bold"),
            bg=self.colors['warning'],
            fg="white",
            command=lambda: self.load_logs(logs_text, today_only=False)
        )
        all_btn.pack(side='left', padx=10, pady=10)
        
        # Logs display frame
        logs_frame = tk.Frame(logs_window, bg=self.colors['card'])
        logs_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Scrollable text widget
        text_frame = tk.Frame(logs_frame)
        text_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        scrollbar_v = ttk.Scrollbar(text_frame)
        scrollbar_v.pack(side='right', fill='y')
        
        scrollbar_h = ttk.Scrollbar(text_frame, orient='horizontal')
        scrollbar_h.pack(side='bottom', fill='x')
        
        logs_text = tk.Text(
            text_frame,
            font=("Courier", 10),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            yscrollcommand=scrollbar_v.set,
            xscrollcommand=scrollbar_h.set,
            wrap='none'
        )
        logs_text.pack(side='left', fill='both', expand=True)
        
        scrollbar_v.config(command=logs_text.yview)
        scrollbar_h.config(command=logs_text.xview)
        
        # Load today's logs by default
        self.load_logs(logs_text, today_only=True)
        
        # Close button
        close_btn = tk.Button(
            logs_window,
            text="Close",
            font=("Arial", 12, "bold"),
            bg=self.colors['primary'],
            fg="white",
            width=20,
            command=logs_window.destroy
        )
        close_btn.pack(pady=20)

    def load_logs(self, text_widget, today_only=True):
        """Load and display gate logs"""
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb[GATE_SHEET] if GATE_SHEET in wb.sheetnames else wb.active
            
            text_widget.config(state='normal')
            text_widget.delete(1.0, tk.END)
            
            today = date.today().strftime("%Y-%m-%d")
            
            # Header
            if today_only:
                text_widget.insert(tk.END, f"📅 Today's Gate Logs ({today})\n\n")
            else:
                text_widget.insert(tk.END, "📋 All Gate Logs\n\n")
                
            text_widget.insert(tk.END, f"{'Name':<15} {'Phone':<12} {'Entry Time':<15} {'Exit Time':<15} {'Status':<8} {'Override':<8}\n")
            text_widget.insert(tk.END, "=" * 90 + "\n")
            
            row_count = 0
            for row in ws.iter_rows(min_row=2):
                name = row[0].value or ""
                phone = row[1].value or ""  # Phone_Number is now column 1
                entry_time = row[2].value or ""  # Entry_Time is now column 2
                exit_time = row[3].value or ""  # Exit_Time is now column 3
                status = row[4].value or ""  # Status is now column 4
                override_used = row[5].value if len(row) > 5 else ""  # Override_Used is now column 5
                
                # Filter for today if requested
                if today_only and entry_time:
                    try:
                        entry_date = datetime.strptime(str(entry_time), "%Y-%m-%d %H:%M:%S").date()
                        if entry_date.strftime("%Y-%m-%d") != today:
                            continue
                    except:
                        continue
                
                # Format times
                entry_display = str(entry_time)[11:19] if entry_time and len(str(entry_time)) > 19 else str(entry_time)
                exit_display = str(exit_time)[11:19] if exit_time and len(str(exit_time)) > 19 else str(exit_time)
                override_display = "YES" if override_used == "YES" else ""
                phone_display = phone[:12] if phone else "N/A"  # Truncate phone for display
                
                text_widget.insert(tk.END, f"{name:<15} {phone_display:<12} {entry_display:<15} {exit_display:<15} {status:<8} {override_display:<8}\n")
                row_count += 1
                
            if row_count == 0:
                text_widget.insert(tk.END, "\nNo records found.\n")
            else:
                text_widget.insert(tk.END, f"\nTotal Records: {row_count}\n")
                
            text_widget.config(state='disabled')
            wb.close()
            
        except Exception as e:
            text_widget.config(state='normal')
            text_widget.delete(1.0, tk.END)
            text_widget.insert(tk.END, f"Error loading logs: {str(e)}")
            text_widget.config(state='disabled')

    def going_out_interface(self):
        """Interface for going out/coming back for work/personal reasons"""
        if not self.known_faces:
            messagebox.showinfo("No Users", "No users registered.")
            return
        
        # Create going out window
        going_out_window = tk.Toplevel(self.root)
        going_out_window.title("Going Out Interface")
        going_out_window.geometry("600x700")
        going_out_window.configure(bg=self.colors['bg'])
        going_out_window.grab_set()
        
        # Title
        title_label = tk.Label(
            going_out_window,
            text="🚶‍♂️ Going Out / Coming Back",
            font=("Arial", 18, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['orange']
        )
        title_label.pack(pady=20)
        
        # Instructions
        instruction_label = tk.Label(
            going_out_window,
            text="Select user and action for temporary office exit/return\n(Separate from main gate entry/exit)",
            font=("Arial", 12),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            justify='center'
        )
        instruction_label.pack(pady=10)
        
        # Main content frame
        content_frame = tk.Frame(going_out_window, bg=self.colors['card'])
        content_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # User selection
        tk.Label(
            content_frame,
            text="Select User:",
            font=("Arial", 14, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text']
        ).pack(pady=(10, 5))
        
        # Listbox for users with current status
        listbox_frame = tk.Frame(content_frame, bg=self.colors['card'])
        listbox_frame.pack(fill='x', padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(listbox_frame)
        scrollbar.pack(side='right', fill='y')
        
        user_listbox = tk.Listbox(
            listbox_frame,
            font=("Arial", 12),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            selectbackground=self.colors['primary'],
            yscrollcommand=scrollbar.set,
            height=6
        )
        user_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=user_listbox.yview)
        
        # Populate with users and their going out status
        user_data = []
        for name in self.known_faces.keys():
            gate_status, _, _ = get_user_status_and_history(name)
            going_out_status, reason_type, going_out_time = get_going_out_status(name)
            
            if gate_status != "INSIDE":
                display_text = f"❌ {name} - Not in office"
                user_data.append((name, "NOT_IN_OFFICE", None, None))
            elif going_out_status == "OUT":
                duration = ""
                if going_out_time:
                    minutes = int((datetime.now() - going_out_time).total_seconds() / 60)
                    duration = f" ({minutes}m)"
                display_text = f"🚶 {name} - Out for {reason_type}{duration}"
                user_data.append((name, "OUT", reason_type, going_out_time))
            else:
                display_text = f"✅ {name} - In office"
                user_data.append((name, "IN_OFFICE", None, None))
            
            user_listbox.insert(tk.END, display_text)
        
        # Reason selection frame
        reason_frame = tk.Frame(content_frame, bg=self.colors['card'])
        reason_frame.pack(fill='x', padx=10, pady=10)
        
        reason_var = tk.StringVar()
        
        reason_title = tk.Label(
            reason_frame,
            text="Select Reason:",
            font=("Arial", 14, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text']
        )
        reason_title.pack(pady=(10, 5))
        
        # Radio buttons for reason
        office_radio = tk.Radiobutton(
            reason_frame,
            text="📊 Office Work",
            variable=reason_var,
            value="Office Work",
            font=("Arial", 12),
            bg=self.colors['card'],
            fg=self.colors['text'],
            selectcolor=self.colors['primary'],
            state='disabled'
        )
        office_radio.pack(pady=2)
        
        personal_radio = tk.Radiobutton(
            reason_frame,
            text="🏠 Personal Work",
            variable=reason_var,
            value="Personal Work",
            font=("Arial", 12),
            bg=self.colors['card'],
            fg=self.colors['text'],
            selectcolor=self.colors['primary'],
            state='disabled'
        )
        personal_radio.pack(pady=2)
        
        # Details entry
        details_label = tk.Label(
            reason_frame,
            text="Additional Details (Optional):",
            font=("Arial", 12),
            bg=self.colors['card'],
            fg=self.colors['text']
        )
        details_label.pack(pady=(10, 5))
        
        details_entry = tk.Entry(
            reason_frame,
            font=("Arial", 12),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            width=40,
            state='disabled'
        )
        details_entry.pack(pady=5)
        
        # Status label
        status_label = tk.Label(
            content_frame,
            text="Please select a user first",
            font=("Arial", 12, "italic"),
            bg=self.colors['card'],
            fg=self.colors['warning']
        )
        status_label.pack(pady=10)
        
        def on_user_select(event=None):
            going_out_window.after_idle(update_selection_state)
        
        def update_selection_state():
            selection = user_listbox.curselection()
            if selection and selection[0] < len(user_data):
                selected_user = user_data[selection[0]]
                user_name, status, current_reason, going_out_time = selected_user
                
                if status == "NOT_IN_OFFICE":
                    office_radio.config(state='disabled')
                    personal_radio.config(state='disabled')
                    details_entry.config(state='disabled')
                    reason_var.set("")
                    details_entry.delete(0, tk.END)
                    status_label.config(
                        text=f"❌ {user_name} is not in office - Cannot process going out",
                        fg=self.colors['danger']
                    )
                elif status == "OUT":
                    office_radio.config(state='disabled')
                    personal_radio.config(state='disabled')
                    details_entry.config(state='disabled')
                    reason_var.set("")
                    details_entry.delete(0, tk.END)
                    
                    duration_text = ""
                    if going_out_time:
                        minutes = int((datetime.now() - going_out_time).total_seconds() / 60)
                        duration_text = f" (out for {minutes} minutes)"
                    
                    status_label.config(
                        text=f"🚶 {user_name} is out for {current_reason}{duration_text} - Ready to mark coming back",
                        fg=self.colors['orange']
                    )
                else:  # IN_OFFICE
                    office_radio.config(state='normal')
                    personal_radio.config(state='normal')
                    details_entry.config(state='normal')
                    reason_var.set("Office Work")
                    details_entry.delete(0, tk.END)
                    status_label.config(
                        text=f"✅ {user_name} is in office - Select reason for going out",
                        fg=self.colors['success']
                    )
            else:
                office_radio.config(state='disabled')
                personal_radio.config(state='disabled')
                details_entry.config(state='disabled')
                reason_var.set("")
                details_entry.delete(0, tk.END)
                status_label.config(
                    text="Please select a user first",
                    fg=self.colors['warning']
                )
        
        user_listbox.bind('<<ListboxSelect>>', on_user_select)
        
        # Action buttons frame
        button_frame = tk.Frame(going_out_window, bg=self.colors['bg'])
        button_frame.pack(pady=20)
        
        def process_going_out():
            selection = user_listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a user from the list.")
                return
            
            selected_user = user_data[selection[0]]
            user_name, status, current_reason, going_out_time = selected_user
            
            if status == "NOT_IN_OFFICE":
                messagebox.showinfo(
                    "User Not Available", 
                    f"{user_name} is not currently in the office.\nPlease ensure they have entered through the main gate first."
                )
                return
            
            try:
                if status == "IN_OFFICE":
                    reason = reason_var.get()
                    if not reason:
                        messagebox.showwarning("No Reason Selected", "Please select a reason for going out.")
                        return
                    
                    details = details_entry.get().strip()
                    confirm_msg = f"Mark {user_name} as going out for:\n{reason}"
                    if details:
                        confirm_msg += f"\nDetails: {details}"
                    
                    if messagebox.askyesno("Confirm Going Out", confirm_msg):
                        result = log_going_out_event(user_name, reason, details)
                        if result == "GOING_OUT":
                            current_time = datetime.now().strftime("%H:%M:%S")
                            message = f"🚶 GOING OUT: {user_name} - {reason} ({current_time})"
                            self.message_label.config(
                                text=message,
                                bg=self.colors['orange'],
                                fg="white"
                            )
                            messagebox.showinfo("Success", f"{user_name} marked as going out for {reason}")
                            going_out_window.destroy()
                            self.root.after(5000, lambda: self.message_label.config(text="", bg=self.colors['card']))
                        
                elif status == "OUT":
                    duration_text = ""
                    if going_out_time:
                        duration = int((datetime.now() - going_out_time).total_seconds() / 60)
                        duration_text = f" (was out for {duration} minutes)"
                    
                    if messagebox.askyesno(
                        "Confirm Coming Back", 
                        f"Mark {user_name} as back from {current_reason}?{duration_text}"
                    ):
                        result = log_going_out_event(user_name, current_reason, "")
                        if result == "COMING_BACK":
                            current_time = datetime.now().strftime("%H:%M:%S")
                            message = f"🏢 COMING BACK: {user_name} - From {current_reason} ({current_time})"
                            self.message_label.config(
                                text=message,
                                bg=self.colors['success'],
                                fg="white"
                            )
                            messagebox.showinfo("Success", f"{user_name} marked as back in office")
                            going_out_window.destroy()
                            self.root.after(5000, lambda: self.message_label.config(text="", bg=self.colors['card']))
                        
            except Exception as e:
                messagebox.showerror("Error", f"Failed to process action: {str(e)}")
        
        action_btn = tk.Button(
            button_frame,
            text="🚶 Process Action",
            font=("Arial", 12, "bold"),
            bg=self.colors['orange'],
            fg="white",
            width=20,
            command=process_going_out
        )
        action_btn.pack(side='left', padx=10)
        
        cancel_btn = tk.Button(
            button_frame,
            text="❌ Cancel",
            font=("Arial", 12, "bold"),
            bg=self.colors['card'],
            fg=self.colors['text'],
            width=15,
            command=going_out_window.destroy
        )
        cancel_btn.pack(side='left', padx=10)

    def view_going_out_logs(self):
        """View going out logs"""
        setup_excel_file()
        
        logs_window = tk.Toplevel(self.root)
        logs_window.title("Going Out Logs")
        logs_window.geometry("1000x600")
        logs_window.configure(bg=self.colors['bg'])
        
        # Title
        title_label = tk.Label(
            logs_window,
            text="🚶 Going Out / Coming Back Logs",
            font=("Arial", 18, "bold"),
            bg=self.colors['bg'],
            fg=self.colors['orange']
        )
        title_label.pack(pady=20)
        
        # Filter frame
        filter_frame = tk.Frame(logs_window, bg=self.colors['card'])
        filter_frame.pack(fill='x', padx=20, pady=10)
        
        # Today's logs button
        today_btn = tk.Button(
            filter_frame,
            text="📅 Today's Logs",
            font=("Arial", 10, "bold"),
            bg=self.colors['primary'],
            fg="white",
            command=lambda: self.load_going_out_logs(logs_text, today_only=True)
        )
        today_btn.pack(side='left', padx=10, pady=10)
        
        # All logs button
        all_btn = tk.Button(
            filter_frame,
            text="📋 All Logs",
            font=("Arial", 10, "bold"),
            bg=self.colors['warning'],
            fg="white",
            command=lambda: self.load_going_out_logs(logs_text, today_only=False)
        )
        all_btn.pack(side='left', padx=10, pady=10)
        
        # Logs display frame
        logs_frame = tk.Frame(logs_window, bg=self.colors['card'])
        logs_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Scrollable text widget
        text_frame = tk.Frame(logs_frame)
        text_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        scrollbar_v = ttk.Scrollbar(text_frame)
        scrollbar_v.pack(side='right', fill='y')
        
        scrollbar_h = ttk.Scrollbar(text_frame, orient='horizontal')
        scrollbar_h.pack(side='bottom', fill='x')
        
        logs_text = tk.Text(
            text_frame,
            font=("Courier", 10),
            bg=self.colors['bg'],
            fg=self.colors['text'],
            yscrollcommand=scrollbar_v.set,
            xscrollcommand=scrollbar_h.set,
            wrap='none'
        )
        logs_text.pack(side='left', fill='both', expand=True)
        
        scrollbar_v.config(command=logs_text.yview)
        scrollbar_h.config(command=logs_text.xview)
        
        # Load today's logs by default
        self.load_going_out_logs(logs_text, today_only=True)
        
        # Close button
        close_btn = tk.Button(
            logs_window,
            text="Close",
            font=("Arial", 12, "bold"),
            bg=self.colors['primary'],
            fg="white",
            width=20,
            command=logs_window.destroy
        )
        close_btn.pack(pady=20)

    def load_going_out_logs(self, text_widget, today_only=True):
        """Load and display going out logs"""
        try:
            wb = load_workbook(EXCEL_FILE)
            if GOING_OUT_SHEET not in wb.sheetnames:
                text_widget.config(state='normal')
                text_widget.delete(1.0, tk.END)
                text_widget.insert(tk.END, "No going out logs found.\n")
                text_widget.config(state='disabled')
                wb.close()
                return
                
            ws_going_out = wb[GOING_OUT_SHEET]
            
            text_widget.config(state='normal')
            text_widget.delete(1.0, tk.END)
            
            today = date.today().strftime("%Y-%m-%d")
            
            # Header
            if today_only:
                text_widget.insert(tk.END, f"📅 Today's Going Out Logs ({today})\n\n")
            else:
                text_widget.insert(tk.END, "📋 All Going Out Logs\n\n")
                
            text_widget.insert(tk.END, f"{'Name':<15} {'Going Out':<20} {'Coming Back':<20} {'Reason':<15} {'Details':<20} {'Status':<10} {'Duration(m)':<12}\n")
            text_widget.insert(tk.END, "=" * 120 + "\n")
            
            row_count = 0
            for row in ws_going_out.iter_rows(min_row=2):
                name = row[0].value or ""
                going_out_time = row[1].value or ""
                coming_back_time = row[2].value or ""
                reason_type = row[3].value or ""
                reason_details = row[4].value or ""
                status = row[5].value or ""
                duration = row[6].value or ""
                
                # Filter for today if requested
                if today_only and going_out_time:
                    try:
                        going_out_date = datetime.strptime(str(going_out_time), "%Y-%m-%d %H:%M:%S").date()
                        if going_out_date.strftime("%Y-%m-%d") != today:
                            continue
                    except:
                        continue
                
                # Format times
                going_out_display = str(going_out_time)[11:19] if going_out_time and len(str(going_out_time)) > 19 else str(going_out_time)
                coming_back_display = str(coming_back_time)[11:19] if coming_back_time and len(str(coming_back_time)) > 19 else str(coming_back_time)
                duration_display = f"{duration}" if duration else ""
                
                # Truncate long details
                details_display = reason_details[:18] + "..." if len(str(reason_details)) > 20 else str(reason_details)
                
                text_widget.insert(tk.END, f"{name:<15} {going_out_display:<20} {coming_back_display:<20} {reason_type:<15} {details_display:<20} {status:<10} {duration_display:<12}\n")
                row_count += 1
                
            if row_count == 0:
                text_widget.insert(tk.END, "\nNo records found.\n")
            else:
                text_widget.insert(tk.END, f"\nTotal Records: {row_count}\n")
                
            text_widget.config(state='disabled')
            wb.close()
            
        except Exception as e:
            text_widget.config(state='normal')
            text_widget.delete(1.0, tk.END)
            text_widget.insert(tk.END, f"Error loading going out logs: {str(e)}")
            text_widget.config(state='disabled')

    def close_application(self):
        """Close the application"""
        if messagebox.askyesno("Exit", "Are you sure you want to exit the Gate Pass System?"):
            if self.is_scanning:
                self.stop_recognition()
            cv2.destroyAllWindows()
            self.root.quit()
            self.root.destroy()
    
    def run(self):
        """Start the GUI application"""
        self.root.protocol("WM_DELETE_WINDOW", self.close_application)
        
        try:
            self.root.mainloop()
        except Exception as e:
            print(f"Error in main loop: {e}")
        finally:
            if self.is_scanning:
                self.stop_recognition()
            cv2.destroyAllWindows()


# ---------------------------
# Main Application Entry Point
# ---------------------------
if __name__ == "__main__":
    app = GatePassGUI()
    app.run()