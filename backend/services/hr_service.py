"""
HR Service Module
Handles business logic for HR operations
"""
from datetime import datetime, date, timedelta
from models import db, Employee, Attendance, Leave, Payroll, JobPosting, LeaveType, LeaveStatus, AttendanceStatus, JobStatus, SalaryType, JobApplication, Interview, Candidate, ApplicationStatus, InterviewStatus
from sqlalchemy import inspect, text
import traceback


def _safe_float(value):
    """Best-effort float conversion that tolerates blanks and bad input."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _safe_date(value):
    """Convert ISO-like strings to date objects when possible."""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).date()
    except (TypeError, ValueError):
        return None


_JOB_APP_CANDIDATE_COLUMN_CHECKED = False


def _ensure_job_application_candidate_column():
    """Ensure job_applications table has candidate_id column, adding if missing."""
    global _JOB_APP_CANDIDATE_COLUMN_CHECKED
    if _JOB_APP_CANDIDATE_COLUMN_CHECKED:
        return True

    try:
        inspector = inspect(db.engine)
        columns = {col['name'] for col in inspector.get_columns('job_applications')}
    except Exception as exc:
        print(f"Warning: unable to inspect job_applications table: {exc}")
        return False

    if 'candidate_id' not in columns:
        try:
            with db.engine.begin() as connection:
                connection.execute(text('ALTER TABLE job_applications ADD COLUMN candidate_id INTEGER'))
        except Exception as exc:
            print(f"Warning: unable to add candidate_id column to job_applications: {exc}")
            return False

    _JOB_APP_CANDIDATE_COLUMN_CHECKED = True
    return True


class HRService:
    """Service class for HR operations"""

    @staticmethod
    def get_dashboard_data():
        """Get HR dashboard summary data"""
        try:
            # Get total employees
            total_employees = Employee.query.count()

            # Get today's attendance
            today = date.today()
            today_attendance = Attendance.query.filter_by(date=today).all()
            present_today = len([a for a in today_attendance if a.status == AttendanceStatus.PRESENT])
            absent_today = len([a for a in today_attendance if a.status == AttendanceStatus.ABSENT])

            # Get employees on leave
            current_leaves = Leave.query.filter(
                Leave.start_date <= today,
                Leave.end_date >= today,
                Leave.status == LeaveStatus.APPROVED
            ).count()

            # Get open job positions
            open_positions = JobPosting.query.filter_by(status=JobStatus.OPEN).count()

            # Get recent activities (last 5)
            recent_activities = []

            # Recent employee joins
            recent_joins = Employee.query.filter(
                Employee.joining_date >= today - timedelta(days=30)
            ).order_by(Employee.joining_date.desc()).limit(3).all()

            for emp in recent_joins:
                recent_activities.append({
                    'type': 'join',
                    'message': f'New Employee Joined: {emp.first_name} {emp.last_name} - {emp.department}',
                    'date': emp.joining_date.isoformat()
                })

            # Recent leave approvals
            recent_leaves = Leave.query.filter(
                Leave.status == LeaveStatus.APPROVED,
                Leave.approved_at >= datetime.utcnow() - timedelta(days=7)
            ).order_by(Leave.approved_at.desc()).limit(2).all()

            for leave in recent_leaves:
                recent_activities.append({
                    'type': 'leave',
                    'message': f'Leave Approved: {leave.employee.first_name} {leave.employee.last_name} - {leave.days_requested} days',
                    'date': leave.approved_at.isoformat()
                })

            # Sort activities by date
            recent_activities.sort(key=lambda x: x['date'], reverse=True)
            recent_activities = recent_activities[:5]

            # Department overview
            departments = db.session.query(
                Employee.department,
                db.func.count(Employee.id).label('count')
            ).group_by(Employee.department).all()

            department_overview = {dept: count for dept, count in departments}

            result = {
                'totalEmployees': total_employees,
                'presentToday': present_today,
                'absentToday': absent_today,
                'onLeave': current_leaves,
                'openPositions': open_positions,
                'recentActivities': recent_activities,
                'departmentOverview': department_overview
            }

            return result

        except Exception as e:
            return {
                'totalEmployees': 0,
                'presentToday': 0,
                'absentToday': 0,
                'onLeave': 0,
                'openPositions': 0,
                'recentActivities': [],
                'departmentOverview': {},
                'error': str(e)
            }

    # Employee Management
    @staticmethod
    def get_employees(department=None, status=None, limit=50):
        """Get all employees with optional filtering"""
        query = Employee.query

        if department:
            query = query.filter_by(department=department)
        if status:
            query = query.filter_by(status=status)

        employees = query.order_by(Employee.created_at.desc()).limit(limit).all()
        return [emp.to_dict() for emp in employees]

    @staticmethod
    def get_employee(employee_id):
        """Get a specific employee by ID"""
        employee = Employee.query.get(employee_id)
        if not employee:
            raise ValueError('Employee not found')
        return employee.to_dict()

    @staticmethod
    def create_employee(employee_data):
        """Create a new employee"""
        required_fields = ['firstName', 'lastName', 'email', 'department', 'designation', 'joiningDate', 'salary']

        for field in required_fields:
            if field not in employee_data:
                raise ValueError(f'Missing required field: {field}')

        # Check if email already exists
        existing = Employee.query.filter_by(email=employee_data['email']).first()
        if existing:
            raise ValueError('Employee with this email already exists')

        # Generate employee ID
        last_employee = Employee.query.order_by(Employee.id.desc()).first()
        employee_id = f"EMP{str(last_employee.id + 1).zfill(4)}" if last_employee else "EMP0001"

        employee = Employee(
            employee_id=employee_id,
            first_name=employee_data['firstName'],
            last_name=employee_data['lastName'],
            email=employee_data['email'],
            phone=employee_data.get('phone'),
            date_of_birth=employee_data.get('dateOfBirth'),
            gender=employee_data.get('gender'),
            address=employee_data.get('address'),
            department=employee_data['department'],
            designation=employee_data['designation'],
            joining_date=datetime.fromisoformat(employee_data['joiningDate']),
            salary=employee_data['salary'],
            salary_type=employee_data.get('salaryType', 'daily'),
            manager_id=employee_data.get('managerId')
        )

        db.session.add(employee)
        db.session.commit()
        return employee.to_dict()

    @staticmethod
    def update_employee(employee_id, employee_data):
        """Update an existing employee"""
        employee = Employee.query.get(employee_id)
        if not employee:
            raise ValueError('Employee not found')

        # Check email uniqueness if being changed
        if 'email' in employee_data and employee_data['email'] != employee.email:
            existing = Employee.query.filter_by(email=employee_data['email']).first()
            if existing:
                raise ValueError('Employee with this email already exists')

        # Update fields
        updatable_fields = {
            'firstName': 'first_name',
            'lastName': 'last_name',
            'email': 'email',
            'phone': 'phone',
            'dateOfBirth': 'date_of_birth',
            'gender': 'gender',
            'address': 'address',
            'department': 'department',
            'designation': 'designation',
            'joiningDate': 'joining_date',
            'salary': 'salary',
            'salaryType': 'salary_type',
            'status': 'status',
            'managerId': 'manager_id'
        }

        for api_field, db_field in updatable_fields.items():
            if api_field in employee_data:
                value = employee_data[api_field]
                if api_field == 'joiningDate' and value:
                    value = datetime.fromisoformat(value)
                elif api_field == 'dateOfBirth' and value:
                    value = datetime.fromisoformat(value).date()
                elif api_field == 'salaryType' and value:
                    value = value.lower()
                setattr(employee, db_field, value)

        db.session.commit()
        return employee.to_dict()

    @staticmethod
    def delete_employee(employee_id):
        """Delete an employee (hard delete with cascade)"""
        employee = Employee.query.get(employee_id)
        if not employee:
            raise ValueError('Employee not found')

        # Delete related records first to avoid foreign key constraints
        # Delete attendances
        Attendance.query.filter_by(employee_id=employee_id).delete()
        # Delete leaves
        Leave.query.filter_by(employee_id=employee_id).delete()
        # Delete payrolls
        Payroll.query.filter_by(employee_id=employee_id).delete()
        # Update subordinates to remove manager reference
        Employee.query.filter_by(manager_id=employee_id).update({'manager_id': None})

        # Finally delete the employee
        db.session.delete(employee)
        db.session.commit()
        return {'message': 'Employee deleted successfully'}

    # Attendance Management
    @staticmethod
    def record_attendance(employee_id, attendance_data):
        """Record attendance for an employee"""
        employee = Employee.query.get(employee_id)
        if not employee:
            raise ValueError('Employee not found')

        attendance_date = datetime.fromisoformat(attendance_data['date']).date()

        # Map status string to enum
        status_map = {
            'present': AttendanceStatus.PRESENT,
            'absent': AttendanceStatus.ABSENT,
            'late': AttendanceStatus.LATE,
            'half_day': AttendanceStatus.HALF_DAY
        }
        status_str = attendance_data.get('status', 'present')
        status = status_map.get(status_str, AttendanceStatus.PRESENT)

        # Check if attendance already exists for this date
        existing = Attendance.query.filter_by(
            employee_id=employee_id,
            date=attendance_date
        ).first()

        if existing:
            # Update existing
            existing.check_in_time = attendance_data.get('checkInTime')
            existing.check_out_time = attendance_data.get('checkOutTime')
            existing.status = status
            existing.hours_worked = attendance_data.get('hoursWorked')
            existing.notes = attendance_data.get('notes')
        else:
            # Create new
            attendance = Attendance(
                employee_id=employee_id,
                date=attendance_date,
                check_in_time=attendance_data.get('checkInTime'),
                check_out_time=attendance_data.get('checkOutTime'),
                status=status,
                hours_worked=attendance_data.get('hoursWorked'),
                notes=attendance_data.get('notes')
            )
            db.session.add(attendance)

        db.session.commit()
        return {'message': 'Attendance recorded successfully'}

    @staticmethod
    def get_employee_attendance(employee_id, start_date=None, end_date=None):
        """Get attendance records for an employee"""
        query = Attendance.query.filter_by(employee_id=employee_id)

        if start_date:
            query = query.filter(Attendance.date >= datetime.fromisoformat(start_date).date())
        if end_date:
            query = query.filter(Attendance.date <= datetime.fromisoformat(end_date).date())

        attendance_records = query.order_by(Attendance.date.desc()).all()
        return [record.to_dict() for record in attendance_records]

    @staticmethod
    def get_all_attendance(start_date=None, end_date=None):
        """Get all attendance records"""
        try:
            query = Attendance.query

            if start_date:
                query = query.filter(Attendance.date >= datetime.fromisoformat(start_date).date())
            if end_date:
                query = query.filter(Attendance.date <= datetime.fromisoformat(end_date).date())

            attendance_records = query.order_by(Attendance.date.desc()).all()
            
            # Convert to dict with error handling for each record
            result = []
            for record in attendance_records:
                try:
                    result.append(record.to_dict())
                except Exception as e:
                    print(f"Error converting attendance record {record.id}: {e}")
                    continue
            
            return result
        except Exception as e:
            print(f"Error fetching attendance records: {e}")
            return []

    @staticmethod
    def get_attendance_summary(start_date=None, end_date=None):
        """Get attendance summary for all employees"""
        if not start_date:
            start_date = date.today() - timedelta(days=30)
        else:
            start_date = datetime.fromisoformat(start_date).date()

        if not end_date:
            end_date = date.today()
        else:
            end_date = datetime.fromisoformat(end_date).date()

        attendance_records = Attendance.query.filter(
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).all()

        summary = {}
        for record in attendance_records:
            emp_id = record.employee_id
            if emp_id not in summary:
                summary[emp_id] = {
                    'employeeName': record.employee.full_name,
                    'totalDays': 0,
                    'presentDays': 0,
                    'absentDays': 0,
                    'lateDays': 0,
                    'halfDays': 0
                }

            summary[emp_id]['totalDays'] += 1
            if record.status == AttendanceStatus.PRESENT:
                summary[emp_id]['presentDays'] += 1
            elif record.status == AttendanceStatus.ABSENT:
                summary[emp_id]['absentDays'] += 1
            elif record.status == AttendanceStatus.LATE:
                summary[emp_id]['lateDays'] += 1
            elif record.status == AttendanceStatus.HALF_DAY:
                summary[emp_id]['halfDays'] += 1

        return list(summary.values())

    # Leave Management
    @staticmethod
    def create_leave_request(employee_id, leave_data):
        """Create a new leave request"""
        employee = Employee.query.get(employee_id)
        if not employee:
            raise ValueError('Employee not found')

        start_date = datetime.fromisoformat(leave_data['startDate']).date()
        end_date = datetime.fromisoformat(leave_data['endDate']).date()

        # Calculate days requested
        days_requested = (end_date - start_date).days + 1

        # Convert leave type to uppercase enum
        leave_type_str = leave_data['leaveType'].upper()

        leave = Leave(
            employee_id=employee_id,
            name=employee.full_name,
            leave_type=LeaveType[leave_type_str],
            start_date=start_date,
            end_date=end_date,
            days_requested=days_requested,
            reason=leave_data.get('reason')
        )

        db.session.add(leave)
        db.session.commit()
        return leave.to_dict()

    @staticmethod
    def get_leave_requests(employee_id=None, status=None):
        """Get leave requests with optional filtering"""
        try:
            query = Leave.query

            if employee_id:
                query = query.filter_by(employee_id=employee_id)
            if status:
                try:
                    query = query.filter_by(status=LeaveStatus(status.upper()))
                except (ValueError, KeyError):
                    # Invalid status, skip filtering
                    pass

            leaves = query.order_by(Leave.created_at.desc()).all()
            
            # Convert to dict with error handling for each record
            result = []
            for leave in leaves:
                try:
                    result.append(leave.to_dict())
                except Exception as e:
                    print(f"Error converting leave record {leave.id}: {e}")
                    continue
            
            return result
        except Exception as e:
            print(f"Error fetching leave requests: {e}")
            return []

    @staticmethod
    def approve_leave_request(leave_id, approved=True, approver_id=None):
        """Approve or reject a leave request"""
        leave = Leave.query.get(leave_id)
        if not leave:
            raise ValueError('Leave request not found')

        if approved:
            leave.status = LeaveStatus.APPROVED
        else:
            leave.status = LeaveStatus.REJECTED

        if approver_id:
            leave.approved_by = approver_id
            leave.approved_at = datetime.utcnow()

        db.session.commit()
        return leave.to_dict()

    @staticmethod
    def get_employee_leave_balance(employee_id):
        """Get leave balance for an employee"""
        # This is a simplified calculation - in real implementation,
        # you'd have leave allocation and tracking logic
        current_year = date.today().year

        # Get approved leaves for current year
        approved_leaves = Leave.query.filter(
            Leave.employee_id == employee_id,
            Leave.status == LeaveStatus.APPROVED,
            Leave.start_date >= date(current_year, 1, 1),
            Leave.end_date <= date(current_year, 12, 31)
        ).all()

        casual_used = sum(leave.days_requested for leave in approved_leaves if leave.leave_type == LeaveType.CASUAL)
        sick_used = sum(leave.days_requested for leave in approved_leaves if leave.leave_type == LeaveType.SICK)
        earned_used = sum(leave.days_requested for leave in approved_leaves if leave.leave_type == LeaveType.EARNED)

        return {
            'employeeId': employee_id,
            'casualLeave': {'used': casual_used, 'total': 12},
            'sickLeave': {'used': sick_used, 'total': 12},
            'earnedLeave': {'used': earned_used, 'total': 21}
        }

    # Payroll Management
    @staticmethod
    def generate_payroll(employee_id, pay_period_data):
        """Generate payroll for an employee based on attendance"""
        employee = Employee.query.get(employee_id)
        if not employee:
            raise ValueError('Employee not found')

        start_date = datetime.fromisoformat(pay_period_data['startDate']).date()
        end_date = datetime.fromisoformat(pay_period_data['endDate']).date()

        salary_value = float(employee.salary)
        total_working_days = pay_period_data.get('totalWorkingDays', 0)
        attended_days = pay_period_data.get('attendedDays', total_working_days)  # Default to full if not provided

        # Calculate base salary based on salary type
        if employee.salary_type == 'hourly':
            # Use total_hours from the request data (frontend sends this)
            total_hours = pay_period_data.get('totalHours', 0)
            base_salary = salary_value * total_hours
            full_monthly_salary = salary_value  # Store the hourly rate itself
        elif employee.salary_type == 'monthly':
            monthly_salary = salary_value
            base_salary = (monthly_salary / total_working_days) * attended_days if total_working_days > 0 and attended_days is not None else monthly_salary
            full_monthly_salary = salary_value
        elif employee.salary_type == 'daily':
            daily_rate = salary_value
            base_salary = daily_rate * attended_days if attended_days is not None else 0
            full_monthly_salary = daily_rate * 30  # Assume 30 days per month
        else:
            # Default to monthly
            monthly_salary = salary_value
            base_salary = (monthly_salary / total_working_days) * attended_days if total_working_days > 0 and attended_days is not None else monthly_salary
            full_monthly_salary = salary_value

        # Get allowances from input (required field)
        allowances = float(pay_period_data.get('allowances', 0))

        # Calculate gross salary and deductions based on salary type
        if employee.salary_type == 'hourly':
            # For hourly employees: gross = hourly_rate * total_hours (no allowances added)
            gross_salary = base_salary
            # Deduction: 5% of gross salary
            deductions = gross_salary * 0.05
        else:
            # For monthly and daily employees: gross = base + allowances
            gross_salary = base_salary + allowances
            
            # Detailed deductions for monthly and daily employees
            # PF: 12% of basic salary (capped at ₹15,000)
            pf_base = min(base_salary, 15000)
            pf_deduction = pf_base * 0.12

            # ESI: 0.75% of gross if gross < ₹21,000
            esi_deduction = gross_salary * 0.0075 if gross_salary < 21000 else 0

            # Professional Tax: ₹200/month (fixed)
            pt_deduction = 200

            # Income Tax: simplified 5% if gross > ₹50,000
            it_deduction = gross_salary * 0.05 if gross_salary > 50000 else 0

            # Total deductions
            deductions = pf_deduction + esi_deduction + pt_deduction + it_deduction

        # Calculate net salary
        net_salary = gross_salary - deductions

        payroll = Payroll(
            employee_id=employee_id,
            name=employee.full_name,
            pay_period_start=start_date,
            pay_period_end=end_date,
            monthly_salary=full_monthly_salary,  # Save full monthly equivalent for display
            salary_type=employee.salary_type,  # Store salary type
            allowances=allowances,
            deductions=deductions,
            gross_salary=gross_salary,
            net_salary=net_salary
        )
        db.session.add(payroll)
        db.session.commit()
        return payroll.to_dict()

    @staticmethod
    def get_employee_payrolls(employee_id):
        """Get payroll records for an employee"""
        payrolls = Payroll.query.filter_by(employee_id=employee_id).order_by(Payroll.pay_period_end.desc()).all()
        return [payroll.to_dict() for payroll in payrolls]

    @staticmethod
    def get_payrolls():
        """Get all payroll records"""
        payrolls = Payroll.query.order_by(Payroll.pay_period_end.desc()).all()
        return [payroll.to_dict() for payroll in payrolls]

    @staticmethod
    def process_payroll(payroll_id):
        """Mark payroll as processed and set payment date"""
        payroll = Payroll.query.get(payroll_id)
        if not payroll:
            raise ValueError('Payroll not found')

        payroll.status = 'processed'
        payroll.payment_date = date.today()
        db.session.commit()
        return payroll.to_dict()

    @staticmethod
    def delete_payroll(payroll_id):
        """Delete a payroll record"""
        payroll = Payroll.query.get(payroll_id)
        if not payroll:
            raise ValueError('Payroll not found')

        db.session.delete(payroll)
        db.session.commit()
        return {'message': 'Payroll deleted successfully'}

    @staticmethod
    def update_payroll(payroll_id, pay_period_data):
        """Update payroll for an employee"""
        payroll = Payroll.query.get(payroll_id)
        if not payroll:
            raise ValueError('Payroll not found')

        employee = payroll.employee
        if not employee:
            raise ValueError('Employee not found for this payroll')

        salary_value = float(employee.salary)
        total_working_days = pay_period_data.get('totalWorkingDays', 0)
        attended_days = pay_period_data.get('attendedDays', total_working_days)

        # Update fields
        start_date = payroll.pay_period_start
        end_date = payroll.pay_period_end
        if 'startDate' in pay_period_data:
            start_date = datetime.fromisoformat(pay_period_data['startDate']).date()
            payroll.pay_period_start = start_date
        if 'endDate' in pay_period_data:
            end_date = datetime.fromisoformat(pay_period_data['endDate']).date()
            payroll.pay_period_end = end_date
        if 'allowances' in pay_period_data:
            payroll.allowances = float(pay_period_data['allowances'])

        # Recalculate base salary based on salary type
        if employee.salary_type == 'hourly':
            # Use total_hours from the request data (frontend sends this)
            total_hours = pay_period_data.get('totalHours', 0)
            base_salary = salary_value * total_hours
            full_monthly_salary = salary_value  # Store the hourly rate itself
        elif employee.salary_type == 'monthly':
            monthly_salary = salary_value
            base_salary = (monthly_salary / total_working_days) * attended_days if total_working_days > 0 and attended_days is not None else monthly_salary
            full_monthly_salary = monthly_salary
        elif employee.salary_type == 'daily':
            daily_rate = salary_value
            base_salary = daily_rate * attended_days if attended_days is not None else 0
            full_monthly_salary = daily_rate * 30  # Assume 30 days per month
        else:
            # Default to monthly
            monthly_salary = salary_value
            base_salary = (monthly_salary / total_working_days) * attended_days if total_working_days > 0 and attended_days is not None else monthly_salary
            full_monthly_salary = monthly_salary

        # Calculate gross salary and deductions based on salary type
        if employee.salary_type == 'hourly':
            # For hourly employees: gross = hourly_rate * total_hours (no allowances added)
            gross_salary = base_salary
            # Deduction: 5% of gross salary
            deductions = gross_salary * 0.05
        else:
            # For monthly and daily employees: gross = base + allowances
            gross_salary = base_salary + payroll.allowances
            
            # Detailed deductions for monthly and daily employees
            # PF: 12% of basic salary (capped at ₹15,000)
            pf_base = min(base_salary, 15000)
            pf_deduction = pf_base * 0.12

            # ESI: 0.75% of gross if gross < ₹21,000
            esi_deduction = gross_salary * 0.0075 if gross_salary < 21000 else 0

            # Professional Tax: ₹200/month (fixed)
            pt_deduction = 200

            # Income Tax: simplified 5% if gross > ₹50,000
            it_deduction = gross_salary * 0.05 if gross_salary > 50000 else 0

            # Total deductions
            deductions = pf_deduction + esi_deduction + pt_deduction + it_deduction

        # Calculate net salary
        net_salary = gross_salary - deductions

        payroll.name = employee.full_name
        payroll.monthly_salary = full_monthly_salary  # Save full monthly equivalent for display
        payroll.salary_type = employee.salary_type  # Update salary type
        payroll.deductions = deductions
        payroll.gross_salary = gross_salary
        payroll.net_salary = net_salary

        db.session.commit()
        return payroll.to_dict()

    @staticmethod
    def generate_payslip(payroll_id):
        """Generate payslip HTML for a payroll record"""
        payroll = Payroll.query.get(payroll_id)
        if not payroll:
            raise ValueError('Payroll not found')

        employee = payroll.employee
        if not employee:
            raise ValueError('Employee not found for this payroll')

        # Format dates
        pay_period_start = payroll.pay_period_start.strftime('%d/%m/%Y') if payroll.pay_period_start else 'N/A'
        pay_period_end = payroll.pay_period_end.strftime('%d/%m/%Y') if payroll.pay_period_end else 'N/A'
        payment_date = payroll.payment_date.strftime('%d/%m/%Y') if payroll.payment_date else 'Pending'

        # Compute actual base salary for the period
        base_salary = payroll.gross_salary - payroll.allowances

        # Compute individual deductions based on salary type
        if payroll.salary_type == 'hourly':
            # For hourly employees: 5% deduction on gross salary
            pf_deduction = 0
            esi_deduction = 0
            pt_deduction = 0
            it_deduction = payroll.gross_salary * 0.05
        else:
            # For monthly and daily employees: detailed deductions
            pf_deduction = min(base_salary, 15000) * 0.12
            esi_deduction = payroll.gross_salary * 0.0075 if payroll.gross_salary < 21000 else 0
            pt_deduction = 200.00
            it_deduction = payroll.gross_salary * 0.05 if payroll.gross_salary > 50000 else 0

        # Determine label for basic earnings row
        if payroll.salary_type == 'hourly':
            basic_label = f'Hourly Rate: ₹{payroll.monthly_salary:.0f}/hr × Hours'
            basic_amount = f"{base_salary:,.2f}"
        else:
            basic_label = 'Monthly Salary'
            basic_amount = f"{payroll.monthly_salary:,.2f}"

        # Generate HTML payslip
        payslip_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Payslip - {employee.full_name}</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 20px;
                    color: #333;
                }}
                .header {{
                    text-align: center;
                    border-bottom: 2px solid #007bff;
                    padding-bottom: 20px;
                    margin-bottom: 30px;
                }}
                .company-name {{
                    font-size: 24px;
                    font-weight: bold;
                    color: #007bff;
                    margin-bottom: 10px;
                }}
                .payslip-title {{
                    font-size: 18px;
                    color: #666;
                }}
                .employee-info {{
                    display: grid;
                    grid-template-columns: 1fr 1fr;
                    gap: 20px;
                    margin-bottom: 30px;
                }}
                .info-section {{
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 5px;
                }}
                .info-section h3 {{
                    margin: 0 0 10px 0;
                    color: #007bff;
                    font-size: 16px;
                }}
                .info-row {{
                    display: flex;
                    justify-content: space-between;
                    margin-bottom: 5px;
                }}
                .salary-breakdown {{
                    margin: 30px 0;
                }}
                .salary-table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 20px 0;
                }}
                .salary-table th, .salary-table td {{
                    padding: 12px;
                    text-align: left;
                    border-bottom: 1px solid #ddd;
                }}
                .salary-table th {{
                    background: #f8f9fa;
                    font-weight: bold;
                }}
                .total-row {{
                    background: #e9ecef;
                    font-weight: bold;
                }}
                .net-salary {{
                    font-size: 18px;
                    color: #28a745;
                    text-align: center;
                    margin: 20px 0;
                    padding: 15px;
                    background: #d4edda;
                    border-radius: 5px;
                }}
                .footer {{
                    margin-top: 40px;
                    text-align: center;
                    color: #666;
                    font-size: 12px;
                }}
                @media print {{
                    body {{ margin: 0; }}
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <div class="company-name">ERP System</div>
                <div class="payslip-title">Employee Payslip</div>
            </div>

            <div class="employee-info">
                <div class="info-section">
                    <h3>Employee Details</h3>
                    <div class="info-row">
                        <span>Name:</span>
                        <span>{employee.full_name}</span>
                    </div>
                    <div class="info-row">
                        <span>Employee ID:</span>
                        <span>{employee.employee_id}</span>
                    </div>
                    <div class="info-row">
                        <span>Department:</span>
                        <span>{employee.department}</span>
                    </div>
                    <div class="info-row">
                        <span>Designation:</span>
                        <span>{employee.designation}</span>
                    </div>
                </div>

                <div class="info-section">
                    <h3>Pay Period</h3>
                    <div class="info-row">
                        <span>From:</span>
                        <span>{pay_period_start}</span>
                    </div>
                    <div class="info-row">
                        <span>To:</span>
                        <span>{pay_period_end}</span>
                    </div>
                    <div class="info-row">
                        <span>Payment Date:</span>
                        <span>{payment_date}</span>
                    </div>
                    <div class="info-row">
                        <span>Status:</span>
                        <span>{payroll.status.title() if payroll.status else 'Pending'}</span>
                    </div>
                </div>
            </div>

            <div class="salary-breakdown">
                <h3 style="color: #007bff; margin-bottom: 20px;">Salary Breakdown</h3>
                <table class="salary-table">
                    <thead>
                        <tr>
                            <th>Description</th>
                            <th>Amount (₹)</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td>{basic_label}</td>
                            <td>{basic_amount}</td>
                        </tr>
                        <tr>
                            <td>Allowances</td>
                            <td>{payroll.allowances:,.2f}</td>
                        </tr>
                        <tr class="total-row">
                            <td>Gross Salary</td>
                            <td>{payroll.gross_salary:,.2f}</td>
                        </tr>
                        <tr>
                            <td colspan="2" style="font-weight: bold; background: #f8f9fa;">Deductions</td>
                        </tr>
                        <tr>
                            <td>&nbsp;&nbsp;Provident Fund (12%)</td>
                            <td>({pf_deduction:,.2f})</td>
                        </tr>
                        <tr>
                            <td>&nbsp;&nbsp;ESI (0.75%)</td>
                            <td>({esi_deduction:,.2f})</td>
                        </tr>
                        <tr>
                            <td>&nbsp;&nbsp;Professional Tax</td>
                            <td>({pt_deduction:,.2f})</td>
                        </tr>
                        <tr>
                            <td>&nbsp;&nbsp;Income Tax (5%)</td>
                            <td>({it_deduction:,.2f})</td>
                        </tr>
                        <tr class="total-row">
                            <td>Total Deductions</td>
                            <td>({payroll.deductions:,.2f})</td>
                        </tr>
                        <tr class="total-row">
                            <td>Net Salary</td>
                            <td>{payroll.net_salary:,.2f}</td>
                        </tr>
                    </tbody>
                </table>
            </div>

            <div class="net-salary">
                <strong>Net Salary Payable: ₹{payroll.net_salary:,.2f}</strong>
            </div>

            <div class="footer">
                <p>This is a computer-generated payslip and does not require a signature.</p>
                <p>Generated on {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
            </div>
        </body>
        </html>
        """

        return payslip_html

    # Recruitment Management
    @staticmethod
    def create_job_posting(job_data):
        """Create a new job posting"""
        required_fields = ['title', 'department', 'description']

        for field in required_fields:
            if field not in job_data:
                raise ValueError(f'Missing required field: {field}')

        job = JobPosting(
            title=job_data['title'],
            department=job_data['department'],
            location=job_data.get('location'),
            employment_type=job_data.get('employmentType'),
            experience_level=job_data.get('experienceLevel'),
            salary_range=job_data.get('salaryRange'),
            description=job_data['description'],
            requirements=job_data.get('requirements'),
            responsibilities=job_data.get('responsibilities'),
            posted_by=job_data.get('postedBy'),
            application_deadline=job_data.get('applicationDeadline')
        )

        if job.application_deadline:
            job.application_deadline = datetime.fromisoformat(job.application_deadline).date()

        db.session.add(job)
        db.session.commit()
        return job.to_dict()

    @staticmethod
    def get_job_postings(status=None):
        """Get job postings with optional status filter"""
        query = JobPosting.query

        if status:
            query = query.filter_by(status=JobStatus(status.upper()))

        jobs = query.order_by(JobPosting.created_at.desc()).all()
        return [job.to_dict() for job in jobs]

    @staticmethod
    def update_job_status(job_id, status):
        """Update job posting status"""
        job = JobPosting.query.get(job_id)
        if not job:
            raise ValueError('Job posting not found')

        job.status = JobStatus(status.upper())
        db.session.commit()
        return job.to_dict()

    @staticmethod
    def update_job_posting(job_id, job_data):
        """Update an existing job posting"""
        job = JobPosting.query.get(job_id)
        if not job:
            raise ValueError('Job posting not found')

        # Update fields
        updatable_fields = {
            'title': 'title',
            'department': 'department',
            'location': 'location',
            'employmentType': 'employment_type',
            'experienceLevel': 'experience_level',
            'salaryRange': 'salary_range',
            'description': 'description',
            'requirements': 'requirements',
            'responsibilities': 'responsibilities',
            'applicationDeadline': 'application_deadline'
        }

        for api_field, db_field in updatable_fields.items():
            if api_field in job_data:
                value = job_data[api_field]
                if api_field == 'applicationDeadline' and value:
                    value = datetime.fromisoformat(value).date()
                setattr(job, db_field, value)
        
        # Update status if provided
        if 'status' in job_data:
            job.status = JobStatus(job_data['status'].upper())

        db.session.commit()
        return job.to_dict()

    @staticmethod
    def delete_job_posting(job_id):
        """Delete a job posting"""
        job = JobPosting.query.get(job_id)
        if not job:
            raise ValueError('Job posting not found')

        db.session.delete(job)
        db.session.commit()
        return {'message': 'Job posting deleted successfully'}

    # Job Application Management
    @staticmethod
    def create_job_application(application_data):
        """Create a new job application along with candidate data when provided."""
        _ensure_job_application_candidate_column()

        required_fields = ['jobPostingId', 'candidateName', 'email', 'phone']
        for field in required_fields:
            if not application_data.get(field):
                raise ValueError(f'Missing required field: {field}')

        job = JobPosting.query.get(application_data['jobPostingId'])
        if not job:
            raise ValueError('Job posting not found')
        if job.status != JobStatus.OPEN:
            raise ValueError('Job posting is not open for applications')

        candidate = Candidate.query.filter_by(email=application_data['email']).first()
        if not candidate:
            candidate_notes = []
            if application_data.get('noticePeriod'):
                candidate_notes.append(f"Notice Period: {application_data['noticePeriod']}")
            if application_data.get('notes'):
                candidate_notes.append(application_data['notes'])

            candidate = Candidate(
                name=application_data['candidateName'],
                email=application_data['email'],
                phone=application_data.get('phone'),
                skills=application_data.get('skills'),
                experience_years=_safe_float(application_data.get('experience')),
                current_company=application_data.get('currentCompany'),
                current_position=application_data.get('currentPosition'),
                location=application_data.get('location'),
                resume_path=application_data.get('resumeUrl'),
                source=application_data.get('source', 'direct'),
                notes="\n".join(candidate_notes) if candidate_notes else None
            )
            db.session.add(candidate)
            db.session.flush()

        existing = JobApplication.query.filter_by(
            job_posting_id=job.id,
            applicant_email=application_data['email']
        ).first()
        if existing:
            raise ValueError('Application already exists for this job and email')

        application = JobApplication(
            job_posting_id=job.id,
            applicant_name=application_data['candidateName'],
            applicant_email=application_data['email'],
            applicant_phone=application_data.get('phone'),
            resume_path=application_data.get('resumeUrl'),
            cover_letter=application_data.get('coverLetter'),
            experience_years=_safe_float(application_data.get('experience')),
            current_salary=_safe_float(application_data.get('currentSalary')),
            expected_salary=_safe_float(application_data.get('expectedSalary')),
            availability_date=_safe_date(application_data.get('availabilityDate')),
            status=ApplicationStatus.SUBMITTED,
            notes=application_data.get('notes'),
            candidate_id=candidate.id if candidate else None
        )

        db.session.add(application)
        db.session.commit()
        return application.to_dict()

    @staticmethod
    def get_job_applications(job_posting_id=None, candidate_id=None, status=None):
        """Get job applications with optional filtering"""
        query = JobApplication.query.options(db.joinedload(JobApplication.candidate))

        if job_posting_id:
            query = query.filter_by(job_posting_id=job_posting_id)
        if candidate_id:
            query = query.filter_by(candidate_id=candidate_id)
        if status:
            try:
                query = query.filter_by(status=ApplicationStatus(status.upper()))
            except (ValueError, KeyError):
                pass

        applications = query.order_by(JobApplication.created_at.desc()).all()
        return [app.to_dict() for app in applications]

    @staticmethod
    def update_application_status(application_id, status, notes=None):
        """Update job application status"""
        application = JobApplication.query.get(application_id)
        if not application:
            raise ValueError('Job application not found')

        application.status = ApplicationStatus(status.upper())
        if notes:
            application.notes = notes

        db.session.commit()
        return application.to_dict()

    @staticmethod
    def delete_job_application(application_id):
        """Delete a job application"""
        application = JobApplication.query.get(application_id)
        if not application:
            raise ValueError('Job application not found')

        db.session.delete(application)
        db.session.commit()
        return {'message': 'Job application deleted successfully'}

    # Interview Management
    @staticmethod
    def schedule_interview(interview_data):
        """Schedule a new interview"""
        required_fields = ['applicationId', 'scheduledDate', 'scheduledTime', 'interviewType']

        for field in required_fields:
            if field not in interview_data:
                raise ValueError(f'Missing required field: {field}')

        # Check if application exists
        application = JobApplication.query.get(interview_data['applicationId'])
        if not application:
            raise ValueError('Job application not found')

        # Parse scheduled time
        from datetime import time
        time_parts = interview_data['scheduledTime'].split(':')
        scheduled_time = time(int(time_parts[0]), int(time_parts[1]))

        interview = Interview(
            job_application_id=interview_data['applicationId'],
            scheduled_date=datetime.fromisoformat(interview_data['scheduledDate']),
            scheduled_time=scheduled_time,
            interview_type=interview_data['interviewType'],
            duration_minutes=interview_data.get('duration_minutes', 60),
            interviewers=interview_data.get('interviewers'),
            location=interview_data.get('location')
        )

        # Update application status to interview_scheduled
        application.status = ApplicationStatus.INTERVIEW_SCHEDULED

        db.session.add(interview)
        db.session.commit()
        return interview.to_dict()

    @staticmethod
    def get_interviews(application_id=None, interviewer_id=None, status=None):
        """Get interviews with optional filtering"""
        query = Interview.query

        if application_id:
            query = query.filter_by(job_application_id=application_id)
        if interviewer_id:
            query = query.filter_by(conducted_by=interviewer_id)
        if status:
            try:
                query = query.filter_by(status=InterviewStatus(status.upper()))
            except (ValueError, KeyError):
                pass

        interviews = query.order_by(Interview.scheduled_date.desc()).all()
        return [interview.to_dict() for interview in interviews]

    @staticmethod
    def update_interview_status(interview_id, status, feedback=None, rating=None, decision=None, interviewer_id=None):
        """Update interview status and results"""
        interview = Interview.query.get(interview_id)
        if not interview:
            raise ValueError('Interview not found')

        interview.status = InterviewStatus(status.upper())
        if feedback:
            interview.feedback = feedback
        if rating is not None:
            interview.rating = rating
        if decision:
            interview.decision = decision
        if interviewer_id:
            interview.conducted_by = interviewer_id
            interview.conducted_at = datetime.utcnow()
        
        # Update application status based on decision
        if decision and interview.job_application:
            if decision == 'selected':
                interview.job_application.status = ApplicationStatus.OFFERED
            elif decision == 'rejected':
                interview.job_application.status = ApplicationStatus.REJECTED

        db.session.commit()
        return interview.to_dict()

    @staticmethod
    def delete_interview(interview_id):
        """Delete an interview"""
        interview = Interview.query.get(interview_id)
        if not interview:
            raise ValueError('Interview not found')

        db.session.delete(interview)
        db.session.commit()
        return {'message': 'Interview deleted successfully'}

    # Candidate Management
    @staticmethod
    def create_candidate(candidate_data):
        """Create a new candidate"""
        required_fields = ['firstName', 'lastName', 'email']

        for field in required_fields:
            if field not in candidate_data:
                raise ValueError(f'Missing required field: {field}')

        # Check if email already exists
        existing = Candidate.query.filter_by(email=candidate_data['email']).first()
        if existing:
            raise ValueError('Candidate with this email already exists')

        candidate = Candidate(
            first_name=candidate_data['firstName'],
            last_name=candidate_data['lastName'],
            email=candidate_data['email'],
            phone=candidate_data.get('phone'),
            address=candidate_data.get('address'),
            resume_url=candidate_data.get('resumeUrl'),
            linkedin_profile=candidate_data.get('linkedinProfile'),
            portfolio_url=candidate_data.get('portfolioUrl'),
            current_company=candidate_data.get('currentCompany'),
            current_position=candidate_data.get('currentPosition'),
            experience_years=candidate_data.get('experienceYears'),
            education=candidate_data.get('education'),
            skills=candidate_data.get('skills'),
            notes=candidate_data.get('notes')
        )

        db.session.add(candidate)
        db.session.commit()
        return candidate.to_dict()

    @staticmethod
    def get_candidates(search=None, status=None):
        """Get candidates with optional filtering"""
        query = Candidate.query

        if search:
            search_term = f"%{search}%"
            query = query.filter(
                db.or_(
                    Candidate.first_name.ilike(search_term),
                    Candidate.last_name.ilike(search_term),
                    Candidate.email.ilike(search_term),
                    Candidate.skills.ilike(search_term)
                )
            )

        candidates = query.order_by(Candidate.created_at.desc()).all()
        return [candidate.to_dict() for candidate in candidates]

    @staticmethod
    def get_candidate(candidate_id):
        """Get a specific candidate by ID"""
        candidate = Candidate.query.get(candidate_id)
        if not candidate:
            raise ValueError('Candidate not found')
        return candidate.to_dict()

    @staticmethod
    def update_candidate(candidate_id, candidate_data):
        """Update an existing candidate"""
        candidate = Candidate.query.get(candidate_id)
        if not candidate:
            raise ValueError('Candidate not found')

        # Check email uniqueness if being changed
        if 'email' in candidate_data and candidate_data['email'] != candidate.email:
            existing = Candidate.query.filter_by(email=candidate_data['email']).first()
            if existing:
                raise ValueError('Candidate with this email already exists')

        # Update fields
        updatable_fields = {
            'firstName': 'first_name',
            'lastName': 'last_name',
            'email': 'email',
            'phone': 'phone',
            'address': 'address',
            'resumeUrl': 'resume_url',
            'linkedinProfile': 'linkedin_profile',
            'portfolioUrl': 'portfolio_url',
            'currentCompany': 'current_company',
            'currentPosition': 'current_position',
            'experienceYears': 'experience_years',
            'education': 'education',
            'skills': 'skills',
            'notes': 'notes'
        }

        for api_field, db_field in updatable_fields.items():
            if api_field in candidate_data:
                setattr(candidate, db_field, candidate_data[api_field])

        db.session.commit()
        return candidate.to_dict()

    @staticmethod
    def delete_candidate(candidate_id):
        """Delete a candidate"""
        candidate = Candidate.query.get(candidate_id)
        if not candidate:
            raise ValueError('Candidate not found')

        # Delete related applications and interviews first
        JobApplication.query.filter_by(candidate_id=candidate_id).delete()
        Interview.query.filter(Interview.application.has(candidate_id=candidate_id)).delete()

        db.session.delete(candidate)
        db.session.commit()
        return {'message': 'Candidate deleted successfully'}

    @staticmethod
    def export_payroll_report():
        """Export payroll report as Excel file"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO

            # Get all payroll records with employee data
            payrolls = Payroll.query.join(Employee).order_by(Payroll.pay_period_end.desc()).all()

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Payroll Report"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center')
            right_alignment = Alignment(horizontal='right')

            # Headers
            headers = [
                'Employee ID', 'Employee Name', 'Department', 'Designation',
                'Pay Period Start', 'Pay Period End', 'Monthly Salary', 'Allowances',
                'Gross Salary', 'Deductions', 'Net Salary', 'Status', 'Payment Date'
            ]

            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Write data
            for row_num, payroll in enumerate(payrolls, 2):
                employee = payroll.employee

                data = [
                    employee.employee_id if employee else '',
                    employee.full_name if employee else '',
                    employee.department if employee else '',
                    employee.designation if employee else '',
                    payroll.pay_period_start.strftime('%d/%m/%Y') if payroll.pay_period_start else '',
                    payroll.pay_period_end.strftime('%d/%m/%Y') if payroll.pay_period_end else '',
                    f"{payroll.monthly_salary:,.2f}" if payroll.monthly_salary else '0.00',
                    f"{payroll.allowances:,.2f}" if payroll.allowances else '0.00',
                    f"{payroll.gross_salary:,.2f}" if payroll.gross_salary else '0.00',
                    f"{payroll.deductions:,.2f}" if payroll.deductions else '0.00',
                    f"{payroll.net_salary:,.2f}" if payroll.net_salary else '0.00',
                    payroll.status.title() if payroll.status else 'Pending',
                    payroll.payment_date.strftime('%d/%m/%Y') if payroll.payment_date else ''
                ]

                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = border
                    if col_num >= 7 and col_num <= 11:  # Salary columns
                        cell.alignment = right_alignment

            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = ws.cell(row=1, column=col_num).column_letter
                max_length = len(header)
                for row in range(1, len(payrolls) + 2):
                    cell_value = str(ws.cell(row=row, column=col_num).value or '')
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 30)


            # (Summary rows removed as per request; only table data will be exported)

            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            return excel_buffer

        except Exception as e:
            print(f"Error generating payroll report: {e}")
            raise ValueError(f"Failed to generate payroll report: {str(e)}")
